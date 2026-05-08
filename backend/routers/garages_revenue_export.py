"""Garage driver-revenue export — Excel download and email report endpoints."""

import os
import logging
from fastapi import APIRouter, HTTPException, Query
from sf_client import sanitize_soql
import cache
from routers.garages_revenue import _compute_revenue

router = APIRouter()
log = logging.getLogger('garages_revenue_export')


def _get_data(tid: str, sd: str, ed: str) -> dict:
    key = f"driver_rev_{tid}_{sd}_{ed}"
    return cache.cached_query_persistent(key, lambda: _compute_revenue(tid, sd, ed), max_stale_hours=26)


# ── Excel ─────────────────────────────────────────────────────────────────────

@router.get("/api/garages/{territory_id}/driver-revenue/export")
def get_driver_revenue_export(
    territory_id: str,
    start_date: str = Query(...),
    end_date:   str = Query(...),
    garage_name: str = Query(''),
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from fastapi.responses import StreamingResponse

    tid = sanitize_soql(territory_id)
    sd  = sanitize_soql(start_date)
    ed  = sanitize_soql(end_date)
    data    = _get_data(tid, sd, ed)
    drivers = data.get('drivers', [])
    summary = data.get('summary', {})

    wb = Workbook()
    hf   = Font(bold=True, color='FFFFFF', size=10)
    hfil = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    hal  = Alignment(horizontal='center')

    def _head(ws, row):
        for c in ws[row]:
            c.font = hf; c.fill = hfil; c.alignment = hal

    def _autowidth(ws):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                (len(str(c.value or '')) for c in col), default=8) + 4

    # Sheet 1: Summary
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = f'Driver Revenue Report — {garage_name or tid}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = f'{start_date} to {end_date}'
    ws.append([])
    ws.append(['Metric', 'Value'])
    _head(ws, 4)
    for row in [
        ['Period Start', start_date],
        ['Period End', end_date],
        ['Tow/Light Revenue (attributed)', summary.get('total_attributed', 0)],
        ['Battery Revenue (attributed)', summary.get('total_battery_revenue', 0)],
        ['Active Drivers', summary.get('total_drivers', 0)],
        ['Total Calls', summary.get('total_calls', 0)],
        ['Note', 'Revenue = SA→WOLI→WO→Total_Amount_Invoiced__c. Battery excluded from Tow/Light.'],
    ]:
        ws.append(row)
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 22

    # Sheet 2: Driver Revenue (by revenue desc)
    ws2 = wb.create_sheet('Driver Revenue')
    ws2.append(['Driver', 'Total Calls', 'Battery Calls', 'Tow/Light Revenue',
                'Battery Revenue', 'Hours Worked', 'Shift Days', 'Rev/Hour (Tow/Light)'])
    _head(ws2, 1)
    for d in sorted(drivers, key=lambda x: -x['revenue']):
        ws2.append([d['name'], d['calls'], d.get('battery_calls', 0),
                    d['revenue'], d.get('battery_revenue', 0),
                    d['hours'], d.get('shift_days', 0), d['rev_per_hour']])
    _autowidth(ws2)

    # Sheet 3: Battery Revenue (by battery_revenue desc)
    ws3 = wb.create_sheet('Battery Revenue')
    ws3.append(['Driver', 'Battery Calls', 'Battery Revenue', 'Avg Rev/Battery Call'])
    _head(ws3, 1)
    for d in sorted(drivers, key=lambda x: -x.get('battery_revenue', 0)):
        if d.get('battery_calls', 0) > 0:
            bc = d['battery_calls']; br = d.get('battery_revenue', 0)
            ws3.append([d['name'], bc, br, round(br / bc, 2) if bc else 0])
    _autowidth(ws3)

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"driver_revenue_{tid}_{sd}_{ed}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


# ── Email ─────────────────────────────────────────────────────────────────────

def _build_email_html(garage_name: str, start_date: str, end_date: str,
                      drivers: list, summary: dict) -> str:
    """Build an HTML email that mirrors the on-screen Revenue tab layout."""
    sorted_by_rev = sorted(drivers, key=lambda x: -x['revenue'])[:20]
    with_hours    = sorted([d for d in drivers if d['hours'] > 0], key=lambda x: -x['rev_per_hour'])[:20]
    batt_drivers  = sorted([d for d in drivers if d.get('battery_calls', 0) > 0],
                            key=lambda x: -x.get('battery_revenue', 0))[:15]

    max_rev  = max((d['revenue'] for d in sorted_by_rev), default=0) or 1
    max_rph  = max((d['rev_per_hour'] for d in with_hours), default=0) or 1
    max_batt = max((d.get('battery_revenue', 0) for d in batt_drivers), default=0) or 1

    def _rph_color(rph):
        return '#10b981' if rph >= 100 else '#f59e0b' if rph >= 60 else '#ef4444'

    def _rev_color(idx):
        return '#3b82f6' if idx < 5 else '#f59e0b' if idx < 15 else '#64748b'

    def _bar_row(name, subtitle, pct, bar_color, val_str):
        pct = max(min(int(pct), 100), 2)
        return (
            f'<tr>'
            f'<td style="width:120px;text-align:right;padding:2px 8px 2px 0;font-size:10px;'
            f'color:#94a3b8;white-space:nowrap;vertical-align:middle">'
            f'{name[:22]}<br><span style="font-size:9px;color:#475569">{subtitle}</span></td>'
            f'<td style="padding:2px 4px;vertical-align:middle">'
            f'<table cellpadding="0" cellspacing="0" border="0" width="100%"><tr>'
            f'<td width="{pct}%" bgcolor="{bar_color}" style="background:{bar_color};'
            f'height:10px;border-radius:2px"> </td><td> </td>'
            f'</tr></table></td>'
            f'<td style="width:72px;text-align:right;padding:2px 4px;font-size:10px;'
            f'color:#e2e8f0;font-weight:bold;white-space:nowrap;vertical-align:middle">{val_str}</td>'
            f'</tr>'
        )

    rev_rows = ''.join(
        _bar_row(d['name'], f"{d['calls']} calls",
                 d['revenue'] / max_rev * 100, _rev_color(i),
                 f"${d['revenue']:,.0f}")
        for i, d in enumerate(sorted_by_rev))

    rph_rows = ''.join(
        _bar_row(d['name'], f"{d['shift_days']}d / {d['hours']}h",
                 d['rev_per_hour'] / max_rph * 100, _rph_color(d['rev_per_hour']),
                 f"${d['rev_per_hour']}/h")
        for d in with_hours) or (
        '<tr><td colspan="3" style="font-size:10px;color:#475569;padding:8px 4px">'
        'No tracked hours for this period.</td></tr>')

    batt_rows = ''.join(
        _bar_row(d['name'], f"{d['battery_calls']} calls",
                 d.get('battery_revenue', 0) / max_batt * 100, '#f59e0b',
                 f"${d.get('battery_revenue', 0):,.0f}")
        for d in batt_drivers)

    batt_section = ''
    if batt_rows:
        batt_total = summary.get('total_battery_revenue', 0)
        batt_section = (
            f'<div style="background:#1e293b;border:1px solid #78350f;border-radius:10px;'
            f'padding:16px;margin-bottom:24px">'
            f'<div style="font-size:12px;font-weight:700;color:#fff;margin-bottom:10px">'
            f'🔋 Battery Revenue per Driver '
            f'<span style="font-size:9px;font-weight:400;color:#64748b">'
            f'· {len(batt_drivers)} drivers · ${batt_total:,.2f} total</span></div>'
            f'<table width="100%" cellpadding="0" cellspacing="2" border="0">{batt_rows}</table>'
            f'<div style="font-size:9px;color:#475569;margin-top:8px;border-top:1px solid #334155;'
            f'padding-top:6px">Battery Jump Start only · Drop-Off excluded</div></div>'
        )

    title = garage_name or 'Garage'
    return (
        f'<div style="font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',Helvetica,Arial,'
        f'sans-serif;max-width:900px;background:#0f172a;color:#e2e8f0;padding:28px;border-radius:12px">'

        f'<h2 style="margin:0 0 4px;color:#fff;font-size:18px">{title} — Driver Revenue Report</h2>'
        f'<p style="margin:0 0 24px;color:#64748b;font-size:12px">{start_date} to {end_date}</p>'

        # Summary cards
        f'<table width="100%" cellpadding="14" cellspacing="8" border="0" style="margin-bottom:28px"><tr>'
        f'<td bgcolor="#1e293b" style="background:#1e293b;border-radius:8px;text-align:center">'
        f'<div style="font-size:9px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Tow/Light Revenue</div>'
        f'<div style="font-size:18px;font-weight:900;color:#34d399">${summary.get("total_attributed",0):,.2f}</div></td>'
        f'<td bgcolor="#1e293b" style="background:#1e293b;border-radius:8px;text-align:center">'
        f'<div style="font-size:9px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Battery Revenue</div>'
        f'<div style="font-size:18px;font-weight:900;color:#fbbf24">${summary.get("total_battery_revenue",0):,.2f}</div></td>'
        f'<td bgcolor="#1e293b" style="background:#1e293b;border-radius:8px;text-align:center">'
        f'<div style="font-size:9px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Active Drivers</div>'
        f'<div style="font-size:18px;font-weight:900;color:#fff">{summary.get("total_drivers",0)}</div></td>'
        f'<td bgcolor="#1e293b" style="background:#1e293b;border-radius:8px;text-align:center">'
        f'<div style="font-size:9px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Total Calls</div>'
        f'<div style="font-size:18px;font-weight:900;color:#fff">{summary.get("total_calls",0):,}</div></td>'
        f'</tr></table>'

        # Charts side by side
        f'<table width="100%" cellpadding="0" cellspacing="12" border="0" style="margin-bottom:24px"><tr valign="top">'

        # Left: Revenue per Driver
        f'<td width="50%" bgcolor="#1e293b" style="background:#1e293b;border-radius:10px;padding:16px">'
        f'<div style="font-size:12px;font-weight:700;color:#fff;margin-bottom:8px">💰 Revenue per Driver'
        f'<span style="font-size:9px;font-weight:400;color:#64748b"> · Top {len(sorted_by_rev)} of {len(drivers)}</span></div>'
        f'<table width="100%" cellpadding="0" cellspacing="2" border="0">{rev_rows}</table>'
        f'<div style="font-size:9px;color:#475569;margin-top:8px;border-top:1px solid #334155;padding-top:6px">'
        f'Tow/Light only · Battery excluded · Blue=top 5 · Amber=6–15</div></td>'

        # Right: Revenue per Hour
        f'<td width="50%" bgcolor="#1e293b" style="background:#1e293b;border-radius:10px;padding:16px">'
        f'<div style="font-size:12px;font-weight:700;color:#fff;margin-bottom:6px">⏱ Revenue per Hour'
        f'<span style="font-size:9px;font-weight:400;color:#64748b"> · {len(with_hours)} tracked</span></div>'
        f'<div style="font-size:9px;color:#64748b;margin-bottom:8px">🟢 ≥$100/h &nbsp; 🟡 $60–$99/h &nbsp; 🔴 &lt;$60/h</div>'
        f'<table width="100%" cellpadding="0" cellspacing="2" border="0">{rph_rows}</table>'
        f'<div style="font-size:9px;color:#475569;margin-top:8px;border-top:1px solid #334155;padding-top:6px">'
        f'Hours = AssetHistory login/logout · sessions capped 16h · open sessions discarded</div></td>'
        f'</tr></table>'

        f'{batt_section}'

        f'<p style="font-size:10px;color:#334155;margin:0;padding-top:16px;border-top:1px solid #1e293b">'
        f'Revenue = SA → WOLI → WO → Total_Amount_Invoiced__c · Battery excluded from Tow/Light · '
        f'Generated by <strong style="color:#6366f1">FleetPulse</strong></p>'
        f'</div>'
    )


@router.post("/api/garages/{territory_id}/driver-revenue/email")
def post_driver_revenue_email(territory_id: str, body: dict):
    import requests as _req

    tid         = sanitize_soql(territory_id)
    to_email    = body.get('to', '')
    start_date  = body.get('start_date', '')
    end_date    = body.get('end_date', '')
    garage_name = body.get('garage_name', '')

    if not to_email or '@' not in to_email:
        raise HTTPException(400, "Valid email required")

    data    = _get_data(tid, sanitize_soql(start_date), sanitize_soql(end_date))
    drivers = data.get('drivers', [])
    summary = data.get('summary', {})

    html    = _build_email_html(garage_name, start_date, end_date, drivers, summary)
    subject = f"{garage_name} Driver Revenue Report — {start_date} to {end_date}"

    agentmail_key   = os.environ.get('AGENTMAIL_API_KEY', '')
    agentmail_inbox = os.environ.get('AGENTMAIL_INBOX', 'fslnyaaa@agentmail.to')
    if not agentmail_key:
        raise HTTPException(500, "Email service not configured (AGENTMAIL_API_KEY)")

    resp = _req.post(
        f"https://api.agentmail.to/v0/inboxes/{agentmail_inbox}/messages/send",
        headers={"Authorization": f"Bearer {agentmail_key}", "Content-Type": "application/json"},
        json={"to": [to_email], "subject": subject, "html": html},
        timeout=15,
    )
    if resp.status_code >= 400:
        log.warning(f"Revenue email send failed: {resp.status_code} {resp.text[:200]}")
        raise HTTPException(500, f"Failed to send email: {resp.text[:200]}")

    return {"status": "sent", "to": to_email}
