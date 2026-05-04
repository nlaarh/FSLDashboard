"""Reporting — batch garage performance summary across multiple garages and a date range."""

import hashlib
import logging
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date as _date, timedelta
from io import BytesIO
from typing import List

from openpyxl.utils import get_column_letter

import cache
import database
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

from sf_batch import batch_soql_parallel
from sf_client import sf_query_all, sf_parallel, sanitize_soql
from routers.garages_scorecard import api_garage_performance_scorecard
from routers.garages_performance import get_performance
from utils import soql_date_range, parse_dt as _parse_dt, is_fleet_territory, totally_satisfied_pct as _sat_pct

router = APIRouter()
log = logging.getLogger('reporting')

# ── Single-garage fallback (used when < 5 garages requested) ──────────────────

def _fetch_garage_row(gid: str, start_date: str, end_date: str) -> dict:
    """Fetch scorecard + performance for one garage (cached per-garage path)."""
    try:
        sc = api_garage_performance_scorecard(gid, start_date, end_date)
        gs = sc.get('garage_summary', {})
        total = gs.get('total_sas', 0) or 0
        completed = gs.get('total_completed', 0) or 0

        try:
            perf = get_performance(gid, period_start=start_date, period_end=end_date)
        except Exception as pex:
            log.warning('Reporting: performance fetch failed for %s: %s', gid, pex)
            perf = {}

        fc = perf.get('first_call') or {}
        rt = perf.get('response_time') or {}
        pts = perf.get('pts_ata') or {}

        return {
            'garage_id': gid,
            'garage_name': '',
            'garage_type': sc.get('garage_type', ''),
            'total_sas': total,
            'completed': completed,
            'completion_pct': round(completed / total * 100, 1) if total else None,
            'declined': gs.get('declined', 0),
            'cancelled': gs.get('cancelled', 0),
            'decline_rate': round(100 * gs.get('declined', 0) / total, 1) if total else None,
            'first_call_pct': fc.get('first_call_pct'),
            'second_call_pct': fc.get('second_call_pct'),
            'accepted_completion_pct': fc.get('accepted_completion_pct'),
            'avg_ata': gs.get('avg_ata'),
            'median_ata': rt.get('median'),
            'under_45_pct': rt.get('under_45_pct'),
            'over_120_pct': rt.get('over_120_pct'),
            'avg_pta': gs.get('avg_pta'),
            'pta_hit_pct': gs.get('pta_hit_pct'),
            'pta_on_time_pct': pts.get('on_time_pct'),
            'pta_avg_delta': pts.get('avg_delta'),
            'total_surveys': gs.get('total_surveys', 0),
            'overall_pct': gs.get('overall_pct'),
            'response_time_pct': gs.get('response_time_pct'),
            'technician_pct': gs.get('technician_pct'),
            'kept_informed_pct': gs.get('kept_informed_pct'),
            'bonus_tier': gs.get('bonus_tier'),
            'bonus_per_sa': gs.get('bonus_per_sa'),
            'total_bonus': gs.get('total_bonus'),
        }
    except Exception as exc:
        log.warning('Reporting: failed to fetch garage %s: %s', gid, exc)
        return {'garage_id': gid, 'garage_name': gid, 'error': str(exc)}


# ── Bulk path (used when ≥ 5 garages requested) ───────────────────────────────

def _accepted(lst):
    return [s for s in lst if not s.get('ERS_Facility_Decline_Reason__c')]


def _compute_bulk_report(territory_ids: list, start_date: str, end_date: str) -> list:
    """3 round-trips total: SAs + surveys in parallel, then SA history batched by ID.
    Groups everything by territory in Python — ~10x faster than per-garage queries."""
    since, until = soql_date_range(start_date, end_date)

    # Territory filter: if ≤ 200 IDs include it; otherwise fetch all and filter in Python
    if len(territory_ids) <= 200:
        tid_csv = ', '.join(f"'{t}'" for t in territory_ids)
        tid_filter = f"AND ServiceTerritoryId IN ({tid_csv})"
    else:
        tid_filter = ''

    # ── Round 1: SAs + surveys in parallel ───────────────────────────────────
    step1 = sf_parallel(
        sas=lambda: sf_query_all(f"""
            SELECT Id, ServiceTerritoryId, ServiceTerritory.Name, Status, CreatedDate,
                   ActualStartTime, ERS_PTA__c, ERS_Dispatch_Method__c,
                   ERS_Facility_Decline_Reason__c, WorkType.Name
            FROM ServiceAppointment
            WHERE CreatedDate >= {since}
              AND CreatedDate < {until}
              AND RecordType.Name = 'ERS Service Appointment'
              AND Status IN ('Dispatched','Completed','Canceled','Assigned',
                             'Cancel Call - Service Not En Route',
                             'Cancel Call - Service En Route',
                             'Unable to Complete','No-Show')
              {tid_filter}
        """),
        surveys=lambda: sf_query_all(f"""
            SELECT ERS_Work_Order__r.ServiceTerritoryId,
                   ERS_Overall_Satisfaction__c,
                   ERS_Response_Time_Satisfaction__c,
                   ERS_Technician_Satisfaction__c,
                   ERSSatisfaction_With_Being_Kept_Informed__c
            FROM Survey_Result__c
            WHERE ERS_Work_Order__r.CreatedDate >= {since}
              AND ERS_Work_Order__r.CreatedDate < {until}
              AND ERS_Overall_Satisfaction__c != null
        """),
    )

    all_sas = step1['sas']
    all_sa_ids = [s['Id'] for s in all_sas]
    log.info('Bulk report: %d SAs, %d surveys fetched', len(all_sas), len(step1['surveys']))

    # ── Round 2: SA history batched by SA ID (parallel chunks of 200) ────────
    # Field='ServiceTerritory' → cascade detection (1st call vs 2nd+ call)
    # Field='Status' → 'On Location' for Towbook real ATA
    all_history = batch_soql_parallel("""
        SELECT ServiceAppointmentId, Field, OldValue, NewValue, CreatedDate
        FROM ServiceAppointmentHistory
        WHERE Field IN ('Status', 'ServiceTerritory')
          AND ServiceAppointmentId IN ('{id_list}')
        ORDER BY ServiceAppointmentId, CreatedDate ASC
    """, all_sa_ids, chunk_size=200) if all_sa_ids else []
    log.info('Bulk report: %d SA history records fetched', len(all_history))

    # ── Build lookup structures ───────────────────────────────────────────────
    towbook_on_loc = {}      # sa_id → earliest On Location datetime
    sa_first_territory = {}  # sa_id → first territory ID (OldValue=null record)

    for h in all_history:
        sa_id = h.get('ServiceAppointmentId')
        if not sa_id:
            continue
        field = h.get('Field', '')
        if field == 'Status' and h.get('NewValue') == 'On Location':
            ts = _parse_dt(h.get('CreatedDate'))
            if ts and (sa_id not in towbook_on_loc or ts < towbook_on_loc[sa_id]):
                towbook_on_loc[sa_id] = ts
        elif field == 'ServiceTerritory' and sa_id not in sa_first_territory:
            if h.get('OldValue') is None:
                nv = h.get('NewValue') or ''
                if len(nv) >= 15 and nv.startswith('0H'):
                    sa_first_territory[sa_id] = nv

    # Survey map: tid → [survey rows]
    territory_surveys = defaultdict(list)
    for sv in step1['surveys']:
        wo = sv.get('ERS_Work_Order__r') or {}
        tid = wo.get('ServiceTerritoryId')
        if tid:
            territory_surveys[tid].append(sv)

    # SA map: tid → [sa rows], excluding Tow Drop-Off
    territory_sas = defaultdict(list)
    territory_names = {}
    for sa in all_sas:
        tid = sa.get('ServiceTerritoryId')
        if not tid:
            continue
        wt = ((sa.get('WorkType') or {}).get('Name') or '').lower()
        if 'drop' in wt:
            continue
        territory_sas[tid].append(sa)
        territory_names[tid] = (sa.get('ServiceTerritory') or {}).get('Name', '')

    # ── Per-territory metrics ─────────────────────────────────────────────────
    requested_set = set(territory_ids)
    rows = []

    for tid, sas in territory_sas.items():
        if tid not in requested_set:
            continue

        name = territory_names.get(tid, tid)
        is_fleet = is_fleet_territory(name)
        total = len(sas)

        completed_sas = [s for s in sas if s.get('Status') == 'Completed']
        declined_cnt = sum(1 for s in sas if s.get('ERS_Facility_Decline_Reason__c'))
        cancelled_cnt = sum(1 for s in sas if 'cancel' in (s.get('Status') or '').lower())

        # 1st call: SA whose first territory assignment was this territory (or no history)
        first_call_sas = [s for s in sas if sa_first_territory.get(s['Id'], tid) == tid]
        second_call_sas = [s for s in sas
                           if sa_first_territory.get(s['Id']) is not None
                           and sa_first_territory[s['Id']] != tid]

        accepted = _accepted(sas)
        accepted_completed = [s for s in accepted if s.get('Status') == 'Completed']

        # ATA + PTA accuracy
        ata_vals = []
        pts_deltas = []
        pta_vals = []

        for sa in completed_sas:
            dm = sa.get('ERS_Dispatch_Method__c') or ''
            created = _parse_dt(sa.get('CreatedDate'))
            actual = towbook_on_loc.get(sa['Id']) if dm == 'Towbook' else _parse_dt(sa.get('ActualStartTime'))

            pta_raw = sa.get('ERS_PTA__c')
            if pta_raw is not None:
                pv = float(pta_raw)
                if 0 < pv < 999:
                    pta_vals.append(pv)

            if created and actual:
                ata = (actual - created).total_seconds() / 60
                if 0 < ata < 480:
                    ata_vals.append(ata)
                    if pta_raw is not None:
                        pv = float(pta_raw)
                        if 0 < pv < 999:
                            expected = created + timedelta(minutes=pv)
                            pts_deltas.append((actual - expected).total_seconds() / 60)

        n_ata = len(ata_vals)
        n_pts = len(pts_deltas)
        on_time = sum(1 for d in pts_deltas if d <= 0)

        # Satisfaction + bonus
        surveys = territory_surveys.get(tid, [])
        tech_pct = _sat_pct(surveys, 'ERS_Technician_Satisfaction__c')
        n_completed = len(completed_sas)

        if is_fleet:
            bonus_tier, bonus_per_sa, total_bonus = 'N/A (Fleet)', 0, 0
        else:
            bonus_per_sa, bonus_tier = database.bonus_for_pct(tech_pct)
            total_bonus = bonus_per_sa * n_completed

        rows.append({
            'garage_id': tid,
            'garage_name': name,
            'garage_type': 'fleet' if is_fleet else 'contractor',
            'total_sas': total,
            'completed': n_completed,
            'completion_pct': round(100 * n_completed / total, 1),
            'declined': declined_cnt,
            'cancelled': cancelled_cnt,
            'decline_rate': round(100 * declined_cnt / total, 1),
            'first_call_pct': round(100 * len(_accepted(first_call_sas)) / len(first_call_sas), 1) if first_call_sas else None,
            'second_call_pct': round(100 * len(_accepted(second_call_sas)) / len(second_call_sas), 1) if second_call_sas else None,
            'accepted_completion_pct': round(100 * len(accepted_completed) / len(accepted), 1) if accepted else None,
            'avg_ata': round(sum(ata_vals) / n_ata) if ata_vals else None,
            'median_ata': round(sorted(ata_vals)[n_ata // 2]) if ata_vals else None,
            'under_45_pct': round(100 * sum(1 for t in ata_vals if t < 45) / n_ata, 1) if ata_vals else None,
            'over_120_pct': round(100 * sum(1 for t in ata_vals if t > 120) / n_ata, 1) if ata_vals else None,
            'avg_pta': round(sum(pta_vals) / len(pta_vals)) if pta_vals else None,
            'pta_hit_pct': round(100 * on_time / n_pts, 1) if pts_deltas else None,
            'pta_on_time_pct': round(100 * on_time / n_pts, 1) if pts_deltas else None,
            'pta_avg_delta': round(sum(pts_deltas) / n_pts, 1) if pts_deltas else None,
            'total_surveys': len(surveys),
            'overall_pct': _sat_pct(surveys, 'ERS_Overall_Satisfaction__c'),
            'response_time_pct': _sat_pct(surveys, 'ERS_Response_Time_Satisfaction__c'),
            'technician_pct': tech_pct,
            'kept_informed_pct': _sat_pct(surveys, 'ERSSatisfaction_With_Being_Kept_Informed__c'),
            'bonus_tier': bonus_tier,
            'bonus_per_sa': bonus_per_sa,
            'total_bonus': total_bonus,
        })

    return sorted(rows, key=lambda r: r.get('garage_name', ''))


# ── API endpoints ─────────────────────────────────────────────────────────────

@router.get('/api/reporting/garage-summary')
def api_reporting_garage_summary(
    garage_ids: List[str] = Query(..., description='Territory IDs'),
    start_date: str = Query(None, description='YYYY-MM-DD'),
    end_date: str = Query(None, description='YYYY-MM-DD'),
):
    """Return one summary row per requested garage for the Reporting table.
    Uses bulk SF queries when ≥ 5 garages are requested (10x faster than per-garage)."""
    today = _date.today()
    if not start_date:
        start_date = today.replace(day=1).isoformat()
    if not end_date:
        end_date = today.isoformat()

    clean_ids = [sanitize_soql(gid) for gid in garage_ids]

    if len(clean_ids) >= 5:
        # Bulk path: 3 SF round-trips for all garages, cached 2 hours
        key_hash = hashlib.md5(','.join(sorted(clean_ids)).encode()).hexdigest()
        cache_key = f'bulk_report_{key_hash}_{start_date}_{end_date}'
        rows = cache.cached_query_persistent(
            cache_key,
            lambda: _compute_bulk_report(clean_ids, start_date, end_date),
            max_stale_hours=2,
        )
    else:
        # Per-garage path: benefits from individual 26-hour scorecard cache
        rows = []
        with ThreadPoolExecutor(max_workers=8) as pool:
            futures = {pool.submit(_fetch_garage_row, gid, start_date, end_date): gid
                       for gid in clean_ids}
            for fut in as_completed(futures):
                rows.append(fut.result())
        rows.sort(key=lambda r: r.get('garage_name', ''))

    return {'rows': rows, 'start_date': start_date, 'end_date': end_date}


@router.get('/api/reporting/garage-summary/export')
def api_reporting_export(
    garage_ids: List[str] = Query(..., description='Territory IDs'),
    start_date: str = Query(None),
    end_date: str = Query(None),
):
    """Export reporting table to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    today = _date.today()
    if not start_date:
        start_date = today.replace(day=1).isoformat()
    if not end_date:
        end_date = today.isoformat()

    summary = api_reporting_garage_summary(garage_ids, start_date, end_date)
    rows = summary['rows']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Garage Performance'

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    center = Alignment(horizontal='center')
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'Garage', 'Type',
        'Total SAs', 'Completed', 'Completion %', 'Declined', 'Cancelled', 'Decline %',
        '1st Call Accept %', '2nd+ Call Accept %', 'Accept Completion %',
        'Avg ATA (min)', 'Median ATA (min)', '% <45 min', '% >2 hr',
        'Avg PTA (min)', 'PTA Hit %', 'PTA On-Time %', 'PTA Avg Delta (min)',
        'Surveys', 'Overall Sat %', 'RT Sat %', 'Tech Sat %', 'KI Sat %',
        'Bonus Tier', 'Bonus/SA ($)', 'Total Bonus ($)',
    ]
    col_keys = [
        'garage_name', 'garage_type',
        'total_sas', 'completed', 'completion_pct', 'declined', 'cancelled', 'decline_rate',
        'first_call_pct', 'second_call_pct', 'accepted_completion_pct',
        'avg_ata', 'median_ata', 'under_45_pct', 'over_120_pct',
        'avg_pta', 'pta_hit_pct', 'pta_on_time_pct', 'pta_avg_delta',
        'total_surveys', 'overall_pct', 'response_time_pct', 'technician_pct', 'kept_informed_pct',
        'bonus_tier', 'bonus_per_sa', 'total_bonus',
    ]
    n_cols = len(headers)

    ws.merge_cells(f'A1:{ws.cell(row=1, column=n_cols).column_letter}1')
    title_cell = ws['A1']
    title_cell.value = f'Garage Performance Report  |  {start_date} → {end_date}'
    title_cell.font = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    amber_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

    def _pct_fill(val):
        if val is None:
            return None
        if val >= 90:
            return green_fill
        if val >= 70:
            return amber_fill
        return red_fill

    pct_keys = ('completion_pct', 'pta_hit_pct', 'pta_on_time_pct', 'overall_pct',
                'response_time_pct', 'technician_pct', 'kept_informed_pct',
                'first_call_pct', 'second_call_pct', 'accepted_completion_pct',
                'under_45_pct', 'over_120_pct')
    center_keys = ('bonus_per_sa', 'total_bonus', 'declined', 'cancelled',
                   'total_sas', 'completed', 'total_surveys', 'avg_ata',
                   'median_ata', 'avg_pta', 'decline_rate', 'pta_avg_delta')

    for row_idx, row in enumerate(rows, 3):
        for col_idx, key in enumerate(col_keys, 1):
            val = row.get(key)
            if isinstance(val, float) and val == int(val):
                val = int(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if key in pct_keys:
                cell.alignment = center
                fill = _pct_fill(row.get(key))
                if fill:
                    cell.fill = fill
            elif key in center_keys:
                cell.alignment = center

    widths = [28, 12, 10, 10, 13, 10, 10, 11, 16, 16, 18, 13, 14, 10, 10, 13, 10, 13, 16, 9, 13, 10, 10, 10, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'garage_report_{start_date}_to_{end_date}.xlsx'
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
