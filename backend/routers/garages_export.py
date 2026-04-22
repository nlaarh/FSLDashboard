"""Garage Performance Scorecard — Excel export and email report endpoints."""

import os
import logging
from datetime import date as _date
from fastapi import APIRouter, HTTPException, Query

from sf_client import sf_query_all, sanitize_soql
from sf_batch import batch_soql_query
from utils import parse_dt as _parse_dt

from routers.garages_scorecard import api_garage_performance_scorecard, api_garage_ai_summary

router = APIRouter()
log = logging.getLogger('garages_export')


@router.get("/api/garages/{territory_id}/performance-scorecard/export")
def api_garage_export(
    territory_id: str,
    start_date: str = Query(None),
    end_date: str = Query(None),
):
    """Export garage performance scorecard to Excel."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse

    territory_id = sanitize_soql(territory_id)
    today = _date.today()
    if not start_date:
        start_date = today.replace(day=1).isoformat()
    if not end_date:
        end_date = today.isoformat()

    # Get scorecard data (will use cache if available)
    scorecard = api_garage_performance_scorecard(territory_id, start_date, end_date)
    gs = scorecard['garage_summary']
    ps = scorecard['primary_vs_secondary']
    drivers = scorecard['drivers']

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    amber_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    def _score_fill(pct):
        if pct is None:
            return None
        if pct >= 92:
            return green_fill
        if pct >= 82:
            return None
        if pct >= 70:
            return amber_fill
        return red_fill

    def _style_header(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = 'Garage Summary'
    ws.append(['Garage Performance Scorecard'])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Period: {start_date} to {end_date}'])
    ws.append([])
    ws.append(['Metric', 'Value'])
    _style_header(ws, 4, 2)
    ws.append(['Total SAs', gs.get('total_sas', 0)])
    ws.append(['Total Completed SAs', gs.get('total_completed', 0)])
    ws.append(['Total Surveys', gs.get('total_surveys', 0)])
    ws.append(['Avg ATA (min)', gs.get('avg_ata')])
    ws.append(['PTA Hit Rate %', gs.get('pta_hit_pct')])
    ws.append([])
    ws.append(['Overall Satisfaction %', gs.get('overall_pct')])
    ws.append(['Response Time Satisfaction %', gs.get('response_time_pct')])
    ws.append(['Technician Satisfaction %', gs.get('technician_pct')])
    ws.append(['Kept Informed Satisfaction %', gs.get('kept_informed_pct')])
    ws.append([])
    ws.append(['Bonus Tier', gs.get('bonus_tier', 'N/A')])
    ws.append(['Bonus per SA', f"${gs.get('bonus_per_sa', 0)}"])
    ws.append(['Total Garage Bonus', f"${gs.get('total_bonus', 0):,}"])
    ws.append(['Bonus Formula', f"Garage Tech {gs.get('technician_pct', '—')}% → ${gs.get('bonus_per_sa', 0)}/SA × {gs.get('total_completed', 0)} completed"])
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 50

    # Color code satisfaction score cells
    for row_num in range(11, 15):
        cell = ws.cell(row=row_num, column=2)
        fill = _score_fill(cell.value)
        if fill:
            cell.fill = fill

    # ── Sheet 2: Driver Breakdown ──
    ws2 = wb.create_sheet('Drivers')
    headers = ['Driver', 'Completed', 'Declined', 'Avg ATA (min)', 'Surveys',
               'Overall %', 'Response Time %', 'Technician %', 'Kept Informed %']
    ws2.append(headers)
    _style_header(ws2, 1, len(headers))

    for d in drivers:
        row = [
            d['name'], d.get('completed', 0), d.get('declined', 0),
            d.get('avg_ata'), d['survey_count'],
            d.get('overall_pct'), d.get('response_time_pct'),
            d.get('technician_pct'), d.get('kept_informed_pct'),
        ]
        ws2.append(row)
        row_num = ws2.max_row
        # Color code scores
        for col_idx, val in [(6, d.get('overall_pct')), (7, d.get('response_time_pct')),
                              (8, d.get('technician_pct')), (9, d.get('kept_informed_pct'))]:
            fill = _score_fill(val)
            if fill:
                ws2.cell(row=row_num, column=col_idx).fill = fill
        # Highlight declined > 0 in red
        if (d.get('declined') or 0) > 0:
            ws2.cell(row=row_num, column=3).fill = red_fill
        # Border all cells
        for col in range(1, len(headers) + 1):
            ws2.cell(row=row_num, column=col).border = thin_border

    # Auto-width
    for col_idx, _ in enumerate(headers, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 16
    ws2.column_dimensions['A'].width = 28

    # ── Sheet 3: Primary vs Secondary ──
    ws3 = wb.create_sheet('Primary vs Secondary')
    ps_headers = ['Metric', 'Overall', 'Primary', 'Secondary']
    ws3.append(ps_headers)
    _style_header(ws3, 1, 4)
    o = ps.get('overall', {})
    p = ps.get('primary', {})
    s = ps.get('secondary', {})
    for label, key in [('Total SAs', 'total_sas'), ('Completed', 'completed'),
                        ('Declined', 'declined'), ('Avg ATA (min)', 'avg_ata'),
                        ('PTA Hit %', 'pta_hit_pct'), ('Surveys', 'survey_count'),
                        ('Overall %', 'overall_pct'), ('Response Time %', 'response_time_pct'),
                        ('Technician %', 'technician_pct'), ('Kept Informed %', 'kept_informed_pct')]:
        ws3.append([label, o.get(key), p.get(key), s.get(key)])
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 14

    # ── Sheet 4: All Survey Details ──
    ws4 = wb.create_sheet('Survey Details')
    sv_headers = ['Driver', 'WO Number', 'Call Date', 'Overall', 'Response Time',
                  'Technician', 'Kept Informed', 'Comment']
    ws4.append(sv_headers)
    _style_header(ws4, 1, len(sv_headers))

    for d in drivers:
        for sv in d.get('surveys', []):
            ws4.append([
                d['name'], sv.get('wo_number', ''), sv.get('call_date', ''),
                sv.get('overall', ''), sv.get('response_time', ''),
                sv.get('technician', ''), sv.get('kept_informed', ''),
                sv.get('comment', ''),
            ])

    ws4.column_dimensions['A'].width = 28
    ws4.column_dimensions['B'].width = 14
    ws4.column_dimensions['C'].width = 12
    for c in ['D', 'E', 'F', 'G']:
        ws4.column_dimensions[c].width = 18
    ws4.column_dimensions['H'].width = 50

    # ── Sheet 5: Flat SA Data (analyst verification) ──
    # Query all SAs with full details + territory history for proof
    start_utc = f"{start_date}T00:00:00Z"
    end_utc = f"{end_date}T23:59:59Z"
    territory_id = sanitize_soql(territory_id)

    sa_detail = sf_query_all(f"""
        SELECT Id, AppointmentNumber, CreatedDate, Status,
               ActualStartTime, ERS_PTA__c, WorkType.Name,
               ERS_Dispatch_Method__c, ERS_Facility_Decline_Reason__c,
               ServiceTerritory.Name, ParentRecordId
        FROM ServiceAppointment
        WHERE ServiceTerritoryId = '{territory_id}'
          AND CreatedDate >= {start_utc}
          AND CreatedDate < {end_utc}
          AND RecordType.Name = 'ERS Service Appointment'
          AND WorkType.Name != 'Tow Drop-Off'
    """)

    # Get territory history for primary/secondary proof
    territory_hist = sf_query_all(f"""
        SELECT ServiceAppointmentId, OldValue, NewValue, CreatedDate
        FROM ServiceAppointmentHistory
        WHERE Field = 'ServiceTerritory'
          AND ServiceAppointment.ServiceTerritoryId = '{territory_id}'
          AND ServiceAppointment.CreatedDate >= {start_utc}
          AND ServiceAppointment.CreatedDate < {end_utc}
          AND ServiceAppointment.RecordType.Name = 'ERS Service Appointment'
        ORDER BY ServiceAppointmentId, CreatedDate ASC
    """)

    # Build original territory map
    original_map = {}  # sa_id -> (original_tid, original_name)
    for row in territory_hist:
        sa_id = row['ServiceAppointmentId']
        if sa_id in original_map:
            continue
        if row.get('OldValue') is None:
            nv = row.get('NewValue') or ''
            if len(nv) >= 15 and nv.startswith('0H'):
                original_map[sa_id] = nv
            elif nv and not nv.startswith('0H'):
                # This is the name row — store as name
                if sa_id not in original_map:
                    original_map[sa_id] = nv  # will be overwritten by ID if ID comes next

    # Build name lookup from territory history
    original_names = {}
    for row in territory_hist:
        sa_id = row['ServiceAppointmentId']
        if sa_id in original_names:
            continue
        if row.get('OldValue') is None:
            nv = row.get('NewValue') or ''
            if nv and not nv.startswith('0H'):
                original_names[sa_id] = nv

    # Get assigned driver for each SA
    sa_id_list = [sa['Id'] for sa in sa_detail[:500]]
    assigned_drivers = {}
    if sa_id_list:
        ar_rows = batch_soql_query("""
                SELECT ServiceAppointmentId, ServiceResource.Name
                FROM AssignedResource
                WHERE ServiceAppointmentId IN ('{id_list}')
            """, sa_id_list, chunk_size=200)
        for ar in ar_rows:
            assigned_drivers[ar['ServiceAppointmentId']] = (ar.get('ServiceResource') or {}).get('Name', '')

    # Build survey lookup by WO Id for joining
    survey_by_wo = {}
    woli_ids = [sa.get('ParentRecordId') for sa in sa_detail if sa.get('ParentRecordId')]
    wo_to_sa_export = {}
    if woli_ids:
        woli_rows = batch_soql_query(
            "SELECT Id, WorkOrderId FROM WorkOrderLineItem WHERE Id IN ('{id_list}')",
            woli_ids, chunk_size=200)
        for r in woli_rows:
            for sa in sa_detail:
                if sa.get('ParentRecordId') == r['Id']:
                    wo_to_sa_export[r['WorkOrderId']] = sa['Id']
                    break
    sa_to_wo = {v: k for k, v in wo_to_sa_export.items()}

    # Get all surveys for this garage/period
    all_surveys = sf_query_all(f"""
        SELECT ERS_Work_Order__c, ERS_Overall_Satisfaction__c,
               ERS_Response_Time_Satisfaction__c, ERS_Technician_Satisfaction__c,
               ERSSatisfaction_With_Being_Kept_Informed__c,
               Customer_Comments__c, ERS_NPS__c, ERS_Renew__c
        FROM Survey_Result__c
        WHERE ERS_Work_Order__r.ServiceTerritoryId = '{territory_id}'
          AND ERS_Work_Order__r.CreatedDate >= {start_utc}
          AND ERS_Work_Order__r.CreatedDate < {end_utc}
          AND ERS_Overall_Satisfaction__c != null
    """)
    # Map WO Id -> survey
    for sv in all_surveys:
        wo_id = sv.get('ERS_Work_Order__c')
        if wo_id:
            sa_id = wo_to_sa_export.get(wo_id)
            if sa_id:
                survey_by_wo[sa_id] = sv

    ws5 = wb.create_sheet('SA Flat Data')
    sa_headers = [
        'SA Number', 'SA Id', 'Created Date', 'Status', 'Work Type',
        'Dispatch Method', 'Current Territory', 'Original Territory',
        'Classification', 'Assigned Driver', 'Decline Reason',
        'ATA (min)', 'PTA (min)', 'PTA Hit',
        'Survey Overall', 'Survey Response Time', 'Survey Technician',
        'Survey Kept Informed', 'NPS', 'Renewal', 'Customer Comment',
    ]
    ws5.append(sa_headers)
    _style_header(ws5, 1, len(sa_headers))

    for sa in sa_detail:
        sa_id = sa['Id']
        created = _parse_dt(sa.get('CreatedDate'))
        actual = _parse_dt(sa.get('ActualStartTime'))
        ata = None
        if created and actual:
            ata_val = (actual - created).total_seconds() / 60
            if 0 < ata_val < 480:
                ata = round(ata_val)
        pta = sa.get('ERS_PTA__c')
        pta_val = float(pta) if pta and 0 < float(pta) < 999 else None
        pta_hit = 'Yes' if ata and pta_val and ata <= pta_val else ('No' if ata and pta_val else '')

        orig_name = original_names.get(sa_id, '')
        orig_id = original_map.get(sa_id, '')
        # Classification
        if orig_id and len(orig_id) >= 15 and orig_id.startswith('0H'):
            classification = 'Primary' if orig_id == territory_id else 'Secondary'
        else:
            classification = 'Primary'  # no history = was always here

        current_territory = (sa.get('ServiceTerritory') or {}).get('Name', '')
        driver = assigned_drivers.get(sa_id, '')

        # Join survey data
        sv = survey_by_wo.get(sa_id, {})

        ws5.append([
            sa.get('AppointmentNumber', ''),
            sa_id,
            created.strftime('%Y-%m-%d %H:%M') if created else '',
            sa.get('Status', ''),
            (sa.get('WorkType') or {}).get('Name', ''),
            sa.get('ERS_Dispatch_Method__c', ''),
            current_territory,
            orig_name or '',
            classification,
            driver,
            sa.get('ERS_Facility_Decline_Reason__c', '') or '',
            ata,
            round(pta_val) if pta_val else '',
            pta_hit,
            sv.get('ERS_Overall_Satisfaction__c', ''),
            sv.get('ERS_Response_Time_Satisfaction__c', ''),
            sv.get('ERS_Technician_Satisfaction__c', ''),
            sv.get('ERSSatisfaction_With_Being_Kept_Informed__c', ''),
            sv.get('ERS_NPS__c', ''),
            sv.get('ERS_Renew__c', ''),
            (sv.get('Customer_Comments__c') or '').strip() if sv else '',
        ])
        row_num = ws5.max_row
        # Color classification
        cls_cell = ws5.cell(row=row_num, column=9)
        if classification == 'Secondary':
            cls_cell.fill = amber_fill
        # Border
        for col in range(1, len(sa_headers) + 1):
            ws5.cell(row=row_num, column=col).border = thin_border

    # Auto-width
    widths = [14, 20, 18, 12, 16, 16, 30, 30, 12, 24, 30, 10, 10, 8,
              20, 20, 20, 20, 6, 8, 50]
    for i, w in enumerate(widths, 1):
        ws5.column_dimensions[ws5.cell(row=1, column=i).column_letter].width = w

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"garage_scorecard_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _build_report_html(garage_name: str, start_date: str, end_date: str, gs: dict, ps: dict, drivers: list, ai_summary: str) -> str:
    """Build rich HTML email body for garage performance report."""
    ov = ps.get('overall', {})

    def sc(pct):
        if pct is None: return '#999'
        if pct >= 92: return '#34d399'
        if pct >= 82: return '#60a5fa'
        if pct >= 70: return '#fbbf24'
        return '#f87171'

    driver_rows = ''
    for d in sorted(drivers, key=lambda x: x.get('survey_count', 0), reverse=True):
        declined_val = d.get('declined', 0)
        declined_color = '#ef4444' if declined_val else '#64748b'
        ata_val = d.get('avg_ata')
        ata_color = '#10b981' if ata_val is not None and ata_val <= 45 else '#ef4444' if ata_val is not None else '#999'
        driver_rows += f"""<tr>
            <td style="padding:6px 10px;border-bottom:1px solid #e2e8f0;font-weight:600">{d['name']}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e2e8f0;text-align:center">{d.get('completed', 0)}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e2e8f0;text-align:center;color:{declined_color};font-weight:700">{declined_val}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e2e8f0;text-align:center;color:{ata_color};font-weight:700">{str(ata_val) + 'm' if ata_val is not None else '—'}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e2e8f0;text-align:center">{d.get('survey_count', 0)}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e2e8f0;text-align:center;color:{sc(d.get('overall_pct'))};font-weight:700">{d.get('overall_pct', '—')}%</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e2e8f0;text-align:center;color:{sc(d.get('response_time_pct'))};font-weight:700">{d.get('response_time_pct', '—')}%</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e2e8f0;text-align:center;color:{sc(d.get('technician_pct'))};font-weight:700">{d.get('technician_pct', '—')}%</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e2e8f0;text-align:center;color:{sc(d.get('kept_informed_pct'))};font-weight:700">{d.get('kept_informed_pct', '—')}%</td>
        </tr>"""

    bonus_bg = '#ecfdf5' if (gs.get('bonus_per_sa') or 0) > 0 else '#f8fafc'
    bonus_border = '#a7f3d0' if (gs.get('bonus_per_sa') or 0) > 0 else '#e2e8f0'
    bonus_color = '#34d399' if (gs.get('bonus_per_sa') or 0) > 0 else '#999'
    ata_color = '#10b981' if (ov.get('avg_ata') or 999) <= 45 else '#ef4444'
    pta_color = '#10b981' if (ov.get('pta_hit_pct') or 0) >= 80 else '#ef4444'

    sat_cells = ''
    for label, pct in [('Overall', gs.get('overall_pct')), ('Response Time', gs.get('response_time_pct')),
                        ('Technician', gs.get('technician_pct')), ('Kept Informed', gs.get('kept_informed_pct'))]:
        val = f"{pct}%" if pct is not None else '—'
        sat_cells += f"""<td style="padding:10px;text-align:center;background:#f8fafc;border:1px solid #e2e8f0">
            <div style="font-size:11px;color:#64748b">{label}</div>
            <div style="font-size:24px;font-weight:800;color:{sc(pct)}">{val}</div>
        </td>"""

    summary_html = (ai_summary or 'No AI summary available.').replace('\n', '<br>')

    return f"""<div style="font-family:Segoe UI,Arial,sans-serif;max-width:700px;color:#1e293b">
        <h2 style="margin:0 0 4px;color:#0f172a">{garage_name}</h2>
        <p style="margin:0 0 20px;color:#64748b;font-size:13px">Performance Report — {start_date} to {end_date}</p>
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:16px;margin-bottom:16px">
            <h3 style="margin:0 0 8px;color:#7c3aed;font-size:14px">Executive Summary</h3>
            <p style="margin:0;font-size:13px;line-height:1.6;color:#334155">{summary_html}</p>
        </div>
        <table style="width:100%;border-collapse:collapse;margin-bottom:16px"><tr>
            <td style="padding:10px;text-align:center;background:#f1f5f9"><div style="font-size:22px;font-weight:800;color:#0f172a">{ov.get('total_sas', 0)}</div><div style="font-size:11px;color:#64748b">Total SAs</div></td>
            <td style="padding:10px;text-align:center;background:#f1f5f9"><div style="font-size:22px;font-weight:800;color:#10b981">{ov.get('completed', 0)}</div><div style="font-size:11px;color:#64748b">Completed</div></td>
            <td style="padding:10px;text-align:center;background:#f1f5f9"><div style="font-size:22px;font-weight:800;color:#ef4444">{ov.get('declined', 0)}</div><div style="font-size:11px;color:#64748b">Declined</div></td>
            <td style="padding:10px;text-align:center;background:#f1f5f9"><div style="font-size:18px;font-weight:700;color:{ata_color}">{str(ov['avg_ata']) + 'm' if ov.get('avg_ata') is not None else '—'}</div><div style="font-size:11px;color:#64748b">Avg ATA</div></td>
            <td style="padding:10px;text-align:center;background:#f1f5f9"><div style="font-size:18px;font-weight:700;color:{pta_color}">{str(ov['pta_hit_pct']) + '%' if ov.get('pta_hit_pct') is not None else '—'}</div><div style="font-size:11px;color:#64748b">PTA Hit Rate</div></td>
        </tr></table>
        <h3 style="margin:0 0 8px;font-size:13px;color:#64748b;text-transform:uppercase;letter-spacing:1px">Satisfaction Scores (Totally Satisfied %)</h3>
        <table style="width:100%;border-collapse:collapse;margin-bottom:16px"><tr>{sat_cells}</tr></table>
        <div style="background:{bonus_bg};border:1px solid {bonus_border};border-radius:8px;padding:12px;margin-bottom:16px">
            <span style="font-size:13px;font-weight:700;color:{bonus_color}">Garage Bonus: Tech {gs.get('technician_pct', '—')}% → ${gs.get('bonus_per_sa', 0)}/SA × {gs.get('total_completed', 0)} completed = ${gs.get('total_bonus', 0)}</span>
        </div>
        <h3 style="margin:0 0 8px;font-size:13px;color:#64748b;text-transform:uppercase;letter-spacing:1px">Driver Breakdown ({len(drivers)} drivers)</h3>
        <table style="width:100%;border-collapse:collapse;font-size:13px;margin-bottom:16px">
            <tr style="background:#1e293b;color:#fff">
                <th style="padding:8px 10px;text-align:left">Driver</th>
                <th style="padding:8px 10px;text-align:center">Completed</th>
                <th style="padding:8px 10px;text-align:center">Declined</th>
                <th style="padding:8px 10px;text-align:center">Avg ATA</th>
                <th style="padding:8px 10px;text-align:center">Surveys</th>
                <th style="padding:8px 10px;text-align:center">Overall</th>
                <th style="padding:8px 10px;text-align:center">Resp Time</th>
                <th style="padding:8px 10px;text-align:center">Technician</th>
                <th style="padding:8px 10px;text-align:center">Informed</th>
            </tr>
            {driver_rows}
        </table>
        <p style="font-size:11px;color:#94a3b8;margin:16px 0 0">Generated by <strong style="color:#6366f1">FleetPulse</strong> · {gs.get('total_surveys', 0)} surveys · {start_date} to {end_date}</p>
        <p style="font-size:10px;color:#64748b;margin:2px 0 0">@NourLaaroubi</p>
    </div>"""


@router.post("/api/garages/{territory_id}/performance-scorecard/email")
def api_garage_email_report(
    territory_id: str,
    body: dict,
):
    """Send garage performance report via email."""
    import requests as _req

    territory_id = sanitize_soql(territory_id)
    to_email = body.get('to', '')
    start_date = body.get('start_date')
    end_date = body.get('end_date')
    garage_name = body.get('garage_name', '')

    if not to_email or '@' not in to_email:
        raise HTTPException(400, "Valid email address required")

    # Get scorecard + AI summary
    scorecard = api_garage_performance_scorecard(territory_id, start_date, end_date)
    gs = scorecard['garage_summary']
    ps = scorecard['primary_vs_secondary']
    drivers = scorecard['drivers']

    # Get AI summary (may already be cached)
    ai_result = api_garage_ai_summary(territory_id, start_date, end_date)
    ai_summary = ai_result.get('summary', '')

    html = _build_report_html(garage_name, start_date, end_date, gs, ps, drivers, ai_summary)
    subject = f"{garage_name} Performance Report — {start_date} to {end_date}"

    # Send via AgentMail
    agentmail_key = os.environ.get('AGENTMAIL_API_KEY', '')
    agentmail_inbox = os.environ.get('AGENTMAIL_INBOX', 'fslnyaaa@agentmail.to')
    if not agentmail_key:
        raise HTTPException(500, "Email service not configured (AGENTMAIL_API_KEY)")

    resp = _req.post(
        f"https://api.agentmail.to/v0/inboxes/{agentmail_inbox}/messages/send",
        headers={"Authorization": f"Bearer {agentmail_key}", "Content-Type": "application/json"},
        json={"to": [to_email], "subject": subject, "html": html},
        timeout=15,
    )
    if resp.status_code >= 400:
        log.warning(f"Email send failed: {resp.status_code} {resp.text}")
        raise HTTPException(500, f"Failed to send email: {resp.text[:200]}")

    return {"status": "sent", "to": to_email}
