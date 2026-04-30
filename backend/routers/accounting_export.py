"""Accounting WOA export — Excel with full fields, formatted table + dashboard tab."""

import tempfile
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from fastapi.responses import FileResponse

from routers.accounting_calc import _SF_BASE

_HDR_FONT  = Font(bold=True, size=11, color='FFFFFF')
_HDR_FILL  = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
_GREEN_FILL = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
_AMBER_FILL = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
_LINK_FONT  = Font(color='0563C1', underline='single', size=10)
_BDR = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
              top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))
_NUM_FMT  = '#,##0.00'
_WRAP_TOP = Alignment(wrap_text=True, vertical='top')
_WRAP_CTR = Alignment(wrap_text=True, vertical='top', horizontal='center')


def _cell(ws, row, col, value, *, fmt=None, fill=None, font=None, align=None, link=None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = _BDR
    if fmt:   c.number_format = fmt
    if fill:  c.fill = fill
    if font:  c.font = font
    if align: c.alignment = align
    if link:  c.hyperlink = link; c.font = _LINK_FONT
    return c


def build_export(items: list, status: str) -> FileResponse:
    wb = Workbook()

    # ── TAB 1: WO Adjustments ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'WO Adjustments'

    HEADERS = [
        '#', 'WOA #', 'Facility', 'WO #', 'Product', 'All Line Items',
        'Description', 'Requested', 'SF Billed', 'Delta', 'Est. USD',
        'Recommendation', 'Confidence', 'Full Reasoning',
        'SF ER Mi (actual)', 'SF ER Mi (est)', 'SF Tow Mi (actual)', 'SF Tow Mi (est)',
        'Long Tow?', 'Long Tow Mi', 'On-Site Min',
        'Vehicle', 'Trouble Code', 'Resolution', 'Coverage',
        'Created', 'Created By',
        'WOA Link', 'WO Link',
    ]
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HDR_FONT; c.fill = _HDR_FILL; c.border = _BDR
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for idx, r in enumerate(items, 2):
        paid = r.get('currently_paid') or 0
        req  = r.get('requested_qty') or 0
        delta = req - paid if paid else req
        rec   = (r.get('recommendation') or '').upper()
        sf    = r.get('sf_miles', {}) if isinstance(r.get('sf_miles'), dict) else {}
        v     = r.get('vehicle', {}) if isinstance(r.get('vehicle'), dict) else {}
        veh   = f'{v.get("make","")} {v.get("model","")} ({v.get("group","")})'.strip(' ()')
        # Full rec_reason — preserve \n so wrap_text renders line breaks in Excel
        full_reason = (r.get('rec_reason') or '').strip()

        col = 1
        _cell(ws, idx, col, idx - 1);                                    col += 1
        _cell(ws, idx, col, r.get('woa_number',''), align=_WRAP_TOP);    col += 1
        _cell(ws, idx, col, r.get('facility',''), align=_WRAP_TOP);      col += 1
        _cell(ws, idx, col, r.get('wo_number',''), align=_WRAP_TOP);     col += 1
        _cell(ws, idx, col, r.get('product','') or 'No WOLI', align=_WRAP_TOP); col += 1
        _cell(ws, idx, col, r.get('woli_summary',''), align=_WRAP_TOP);  col += 1
        # Description — full text with line breaks
        _cell(ws, idx, col, (r.get('description') or '').replace('\\n', '\n'), align=_WRAP_TOP); col += 1
        _cell(ws, idx, col, req,   fmt=_NUM_FMT, align=_WRAP_TOP);       col += 1
        _cell(ws, idx, col, paid,  fmt=_NUM_FMT, align=_WRAP_TOP);       col += 1
        _cell(ws, idx, col, round(delta, 2), fmt=_NUM_FMT, align=_WRAP_TOP); col += 1
        est = r.get('estimated_usd')
        _cell(ws, idx, col, round(est, 2) if est is not None else None, fmt=_NUM_FMT, align=_WRAP_TOP); col += 1
        rec_label = 'Approve' if rec == 'APPROVE' else 'Review'
        _cell(ws, idx, col, rec_label,
              fill=_GREEN_FILL if rec == 'APPROVE' else _AMBER_FILL, align=_WRAP_CTR); col += 1
        _cell(ws, idx, col, r.get('confidence',''), align=_WRAP_CTR);    col += 1
        # Full reasoning with all \n steps
        _cell(ws, idx, col, full_reason, align=_WRAP_TOP);               col += 1
        _cell(ws, idx, col, sf.get('enroute'), fmt=_NUM_FMT, align=_WRAP_TOP);         col += 1
        _cell(ws, idx, col, sf.get('estimated_enroute'), fmt=_NUM_FMT, align=_WRAP_TOP); col += 1
        _cell(ws, idx, col, sf.get('tow'), fmt=_NUM_FMT, align=_WRAP_TOP);             col += 1
        _cell(ws, idx, col, sf.get('estimated_tow'), fmt=_NUM_FMT, align=_WRAP_TOP);   col += 1
        _cell(ws, idx, col, 'Yes' if r.get('long_tow_used') else 'No', align=_WRAP_CTR); col += 1
        _cell(ws, idx, col, r.get('long_tow_miles'), fmt=_NUM_FMT, align=_WRAP_TOP);   col += 1
        # on_location_minutes not in list items — placeholder
        _cell(ws, idx, col, None, align=_WRAP_TOP);                       col += 1
        _cell(ws, idx, col, veh if veh != '()' else '', align=_WRAP_TOP); col += 1
        _cell(ws, idx, col, None, align=_WRAP_TOP);  # trouble_code not in list col += 1; col += 1
        _cell(ws, idx, col, None, align=_WRAP_TOP);  col += 1  # resolution
        _cell(ws, idx, col, None, align=_WRAP_TOP);  col += 1  # coverage
        _cell(ws, idx, col, r.get('created_date',''), align=_WRAP_TOP);  col += 1
        _cell(ws, idx, col, r.get('created_by',''), align=_WRAP_TOP);    col += 1
        woa_url = f'{_SF_BASE}/{r.get("id","")}'
        _cell(ws, idx, col, woa_url, link=woa_url, align=_WRAP_TOP);     col += 1
        wo_url  = f'{_SF_BASE}/{r.get("wo_id","")}'
        _cell(ws, idx, col, wo_url, link=wo_url, align=_WRAP_TOP)

    # Column widths — description + reasoning get wide columns for text
    widths = [4, 13, 28, 12, 26, 30,
              40, 10, 10, 10, 12,
              13, 12, 60,
              14, 14, 14, 14,
              10, 10, 12,
              22, 13, 13, 20,
              13, 18,
              14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Row height: let Excel auto-size via wrap_text; set a minimum
    for row in range(2, len(items) + 2):
        ws.row_dimensions[row].height = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{len(items) + 1}'

    # ── TAB 2: Dashboard ──────────────────────────────────────────────────────
    ds = wb.create_sheet('Dashboard')
    title_font   = Font(bold=True, size=14, color='1E293B')
    section_font = Font(bold=True, size=12, color='334155')
    kpi_font     = Font(bold=True, size=18, color='1E293B')
    kpi_label    = Font(size=10, color='64748B')
    kpi_fill     = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    total     = len(items)
    approves  = sum(1 for r in items if r.get('recommendation') == 'approve')
    reviews   = total - approves
    total_req = sum(r.get('requested_qty') or 0 for r in items)
    total_paid= sum(r.get('currently_paid') or 0 for r in items if r.get('currently_paid'))
    total_est = sum(r.get('estimated_usd') or 0 for r in items if r.get('estimated_usd'))
    credits   = sum(1 for r in items if (r.get('requested_qty') or 0) < 0)

    prod_counts, prod_approve = Counter(), Counter()
    prod_req, prod_est = defaultdict(float), defaultdict(float)
    fac_counts, fac_req = Counter(), defaultdict(float)
    by_counts, by_req   = Counter(), defaultdict(float)
    for r in items:
        p  = (r.get('product') or '--').split(' - ')[0].strip() if ' - ' in (r.get('product') or '') else (r.get('product') or '--')
        f  = r.get('facility') or 'Unknown'
        by = r.get('created_by') or 'Unknown'
        prod_counts[p] += 1
        if r.get('recommendation') == 'approve': prod_approve[p] += 1
        prod_req[p]  += r.get('requested_qty') or 0
        prod_est[p]  += r.get('estimated_usd') or 0
        fac_counts[f] += 1; fac_req[f] += r.get('requested_qty') or 0
        by_counts[by] += 1; by_req[by] += r.get('requested_qty') or 0

    ds.cell(row=1, column=1, value='WO Adjustment Dashboard').font = title_font
    ds.merge_cells('A1:H1')
    kpi_data = [
        ('Total WOAs', total), ('Approve', approves), ('Review', reviews),
        ('Approve %', f'{approves * 100 // total}%' if total else '0%'),
        ('Credits', credits), ('Total Requested', round(total_req, 2)),
        ('SF Billed', round(total_paid, 2)), ('Est. USD Exposure', round(total_est, 2)),
    ]
    for i, (label, val) in enumerate(kpi_data):
        col = i + 1
        c = ds.cell(row=3, column=col, value=val)
        c.font = kpi_font; c.fill = kpi_fill
        c.alignment = Alignment(horizontal='center'); c.border = _BDR
        if isinstance(val, float): c.number_format = _NUM_FMT
        lc = ds.cell(row=4, column=col, value=label)
        lc.font = kpi_label; lc.alignment = Alignment(horizontal='center')
        ds.column_dimensions[get_column_letter(col)].width = 16

    def _section(title, start_row, col_headers, data_rows):
        ds.cell(row=start_row, column=1, value=title).font = section_font
        r = start_row + 1
        for i, h in enumerate(col_headers, 1):
            c = ds.cell(row=r, column=i, value=h); c.font = _HDR_FONT; c.fill = _HDR_FILL; c.border = _BDR
        for vals in data_rows:
            r += 1
            for i, v in enumerate(vals, 1):
                c = ds.cell(row=r, column=i, value=v); c.border = _BDR
                if isinstance(v, float): c.number_format = _NUM_FMT
        return r

    prod_rows = [(code, cnt, prod_approve.get(code,0), cnt - prod_approve.get(code,0),
                  round(prod_approve.get(code,0)/cnt,2) if cnt else 0,
                  round(prod_req.get(code,0),2), round(prod_est.get(code,0),2))
                 for code, cnt in prod_counts.most_common()]
    prod_end_row = _section('By Product', 6,
        ['Product','Count','Approve','Review','Approve %','Total Requested','Est. USD'],
        prod_rows)

    chart1 = BarChart()
    chart1.type = 'col'; chart1.title = 'WOAs by Product'; chart1.style = 10
    chart1.width = 20; chart1.height = 12
    chart1.add_data(Reference(ds, min_col=3, min_row=8, max_row=prod_end_row), titles_from_data=True)
    chart1.add_data(Reference(ds, min_col=4, min_row=8, max_row=prod_end_row), titles_from_data=True)
    chart1.set_categories(Reference(ds, min_col=1, min_row=9, max_row=prod_end_row))
    chart1.series[0].graphicalProperties.solidFill = '22C55E'
    chart1.series[1].graphicalProperties.solidFill = 'F59E0B'
    ds.add_chart(chart1, 'I6')

    fac_rows = [(f, cnt, round(fac_req.get(f,0),2)) for f, cnt in fac_counts.most_common(15)]
    fac_end_row = _section('Top 15 Facilities', prod_end_row + 3, ['Facility','Count','Total Requested'], fac_rows)

    by_rows = [(n, cnt, round(by_req.get(n,0),2)) for n, cnt in by_counts.most_common(15)]
    _section('Submitted By', fac_end_row + 3, ['Created By','Count','Total Requested'], by_rows)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return FileResponse(
        path=tmp.name,
        filename=f'WO_Adjustments_{status}.xlsx',
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Cache-Control': 'no-cache, no-store, must-revalidate'},
    )
