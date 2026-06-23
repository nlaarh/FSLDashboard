"""Generic admin CRUD for reference tables stored in Postgres core schema."""
import os
import io
import logging
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
import openpyxl
import db_adapter
from routers.accounting_calc import invalidate_hdv_cache as _invalidate_hdv_cache, invalidate_fuel_cache as _invalidate_fuel_cache

log = logging.getLogger("admin_reference")
router = APIRouter()

_ADMIN_PIN = os.getenv("ADMIN_PIN")
_HDV_TABLE_KEY = 'heavy_duty_vehicles'
_FUEL_TABLE_KEY = 'fuel_reimbursement'
_REFERENCE_ROLES = {'executive', 'ers-director'}


def _maybe_invalidate_caches(table_key: str) -> None:
    """Invalidate relevant caches whenever a reference table is modified."""
    if table_key == _HDV_TABLE_KEY:
        _invalidate_hdv_cache()
    elif table_key == _FUEL_TABLE_KEY:
        _invalidate_fuel_cache()


def _access_check(request: Request, pin: str = "") -> None:
    """Allow reference-data roles via session cookie; fall back to PIN for admin/superadmin."""
    from routers.auth import _verify_cookie
    import users as _users
    cookie = request.cookies.get("fslapp_auth")
    payload = _verify_cookie(cookie) if cookie else None
    if payload:
        username = payload.split(":")[0]
        role = (_users.get_user(username) or {}).get("role", "")
        if role in _REFERENCE_ROLES:
            return
    if not _ADMIN_PIN:
        raise HTTPException(status_code=503, detail="ADMIN_PIN not configured")
    pin_val = request.headers.get("X-Admin-Pin", "") or pin
    if pin_val != _ADMIN_PIN:
        raise HTTPException(status_code=403, detail="Invalid PIN")


# ── Registry ──────────────────────────────────────────────────────────────────
REGISTRY = {
    "accounting_rates": {
        "label": "Accounting Reference Rates",
        "table": "accounting_rates",
        "pk": "code",
        "pk_type": "text",
        "columns": [
            {"key": "code",     "label": "Code",     "type": "text",   "required": True, "readonly": True},
            {"key": "label",    "label": "Label",    "type": "text",   "required": True},
            {"key": "category", "label": "Category", "type": "text"},
            {"key": "value",    "label": "Value",    "type": "number"},
            {"key": "unit",     "label": "Unit",     "type": "text"},
            {"key": "notes",    "label": "Notes",    "type": "text"},
        ],
        "order_by": "category, code",
    },
    "bonus_tiers": {
        "label": "Contractor Bonus Tiers",
        "table": "bonus_tiers",
        "pk": "id",
        "pk_type": "int",
        "columns": [
            {"key": "id",          "label": "ID",       "type": "number", "readonly": True},
            {"key": "min_pct",     "label": "Min %",    "type": "number", "required": True},
            {"key": "bonus_per_sa","label": "Bonus/SA", "type": "number", "required": True},
            {"key": "label",       "label": "Label",    "type": "text"},
            {"key": "sort_order",  "label": "Order",    "type": "number"},
        ],
        "order_by": "min_pct DESC",
    },
    "heavy_duty_vehicles": {
        "label": "Heavy Duty Vehicle List",
        "table": "ref_heavy_duty_vehicles",
        "pk": "id",
        "pk_type": "int",
        "columns": [
            {"key": "id",       "label": "ID",       "type": "number",  "readonly": True},
            {"key": "make",     "label": "Make",     "type": "text",    "required": True},
            {"key": "model",    "label": "Model",    "type": "text",    "required": True},
            {"key": "approved", "label": "Approved", "type": "boolean", "default": True},
            {"key": "notes",    "label": "Notes",    "type": "text"},
        ],
        "order_by": "make, model",
    },
    "fuel_reimbursement": {
        "label": "Max Fuel Reimbursement Allowance",
        "table": "ref_fuel_reimbursement",
        "pk": "dispatch_code",
        "pk_type": "text",
        "columns": [
            {"key": "dispatch_code", "label": "Dispatch Code", "type": "text",   "required": True, "readonly": True},
            {"key": "fuel_type",     "label": "Fuel Type",     "type": "text",   "required": True},
            {"key": "amount_usd",    "label": "Max Amount ($)", "type": "number", "required": True},
        ],
        "order_by": "dispatch_code",
    },
}


def _get_reg(table_key: str) -> dict:
    reg = REGISTRY.get(table_key)
    if not reg:
        raise HTTPException(status_code=404, detail=f"Unknown table: {table_key}")
    return reg


@router.get("/api/admin/reference/tables")
def list_tables(request: Request, pin: str = ""):
    _access_check(request, pin)
    return [{"key": k, "label": v["label"]} for k, v in REGISTRY.items()]


@router.get("/api/admin/reference/{table_key}")
def get_rows(table_key: str, request: Request, pin: str = ""):
    _access_check(request, pin)
    reg = _get_reg(table_key)
    with db_adapter.reader() as db:
        db.execute(f"SELECT * FROM {reg['table']} ORDER BY {reg['order_by']}")
        rows = db.fetchall()
    return {"columns": reg["columns"], "rows": rows, "pk": reg["pk"]}


@router.post("/api/admin/reference/{table_key}")
async def add_row(table_key: str, request: Request):
    _access_check(request)
    reg = _get_reg(table_key)
    body = await request.json()
    editable = [c for c in reg["columns"] if not c.get("readonly")]
    cols = [c["key"] for c in editable]
    # Apply registry default when field is absent from body (prevents NOT NULL violations)
    values = [body.get(c["key"], c.get("default")) for c in editable]
    placeholders = ", ".join(["%s"] * len(cols))
    col_names = ", ".join(cols)
    with db_adapter.writer() as db:
        db.execute(
            f"INSERT INTO {reg['table']} ({col_names}) VALUES ({placeholders})",
            values,
        )
    _maybe_invalidate_caches(table_key)
    return {"ok": True}


# NOTE: Export and import routes must be declared BEFORE /{pk_val} routes to
# prevent FastAPI from matching "export"/"import" as pk_val path parameters.

@router.get("/api/admin/reference/{table_key}/export")
def export_table(table_key: str, request: Request, pin: str = ""):
    """Export table to Excel. Accepts pin from header OR ?pin= query param for browser downloads."""
    _access_check(request, pin)
    reg = _get_reg(table_key)
    with db_adapter.reader() as db:
        db.execute(f"SELECT * FROM {reg['table']} ORDER BY {reg['order_by']}")
        rows = db.fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    col_keys = [c["key"] for c in reg["columns"]]
    col_labels = [c["label"] for c in reg["columns"]]
    ws.append(col_labels)
    for row in rows:
        ws.append([row.get(k) for k in col_keys])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={table_key}.xlsx"},
    )


@router.post("/api/admin/reference/{table_key}/import")
async def import_table(table_key: str, request: Request, file: UploadFile = File(...)):
    _access_check(request)
    reg = _get_reg(table_key)
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    # Map Excel header labels → column keys
    label_to_key = {c["label"]: c["key"] for c in reg["columns"]}
    col_map = {i: label_to_key[h] for i, h in enumerate(headers) if h in label_to_key}

    editable = [c["key"] for c in reg["columns"] if not c.get("readonly")]
    inserted = 0
    with db_adapter.writer() as db:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            mapped = {col_map[i]: v for i, v in enumerate(row) if i in col_map}
            cols = [c for c in editable if c in mapped]
            if not cols:
                continue
            vals = [mapped[c] for c in cols]
            col_names = ", ".join(cols)
            placeholders = ", ".join(["%s"] * len(cols))
            # Use ON CONFLICT DO NOTHING to survive duplicate imports cleanly
            db.execute(
                f"INSERT INTO {reg['table']} ({col_names}) VALUES ({placeholders}) "
                f"ON CONFLICT DO NOTHING",
                vals,
            )
            inserted += 1
    _maybe_invalidate_caches(table_key)
    return {"ok": True, "imported": inserted}


@router.put("/api/admin/reference/{table_key}/{pk_val}")
async def update_row(table_key: str, pk_val: str, request: Request):
    _access_check(request)
    reg = _get_reg(table_key)
    body = await request.json()
    editable = [c["key"] for c in reg["columns"] if not c.get("readonly")]
    sets = ", ".join([f"{c} = %s" for c in editable])
    values = [body.get(c) for c in editable]
    pk_typed = int(pk_val) if reg["pk_type"] == "int" else pk_val
    with db_adapter.writer() as db:
        db.execute(
            f"UPDATE {reg['table']} SET {sets} WHERE {reg['pk']} = %s",
            values + [pk_typed],
        )
    _maybe_invalidate_caches(table_key)
    return {"ok": True}


@router.delete("/api/admin/reference/{table_key}/{pk_val}")
def delete_row(table_key: str, pk_val: str, request: Request):
    _access_check(request)
    reg = _get_reg(table_key)
    pk_typed = int(pk_val) if reg["pk_type"] == "int" else pk_val
    with db_adapter.writer() as db:
        db.execute(f"DELETE FROM {reg['table']} WHERE {reg['pk']} = %s", (pk_typed,))
    _maybe_invalidate_caches(table_key)
    return {"ok": True}
