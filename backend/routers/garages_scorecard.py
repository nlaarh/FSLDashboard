"""Garage Performance Scorecard — satisfaction scores, bonus calculation, driver breakdown."""

import os
import json as _json
import logging
import requests as _requests
from datetime import date as _date, timedelta
from collections import defaultdict
from fastapi import APIRouter, HTTPException, Query

from sf_client import sf_query_all, sf_parallel, sanitize_soql
from utils import parse_dt as _parse_dt
import cache

router = APIRouter()
log = logging.getLogger('garages_scorecard')

_SETTINGS_FILE = os.path.expanduser('~/.fslapp/settings.json')

# Bonus tiers based on Technician "Totally Satisfied" %
BONUS_TIERS = [
    (98, 4),   # ≥ 98% → $4/SA
    (96, 3),   # 96-97.99% → $3/SA
    (94, 2),   # 94-95.99% → $2/SA
    (92, 1),   # 92-93.99% → $1/SA
]


def _bonus_for_pct(pct):
    """Return (bonus_per_sa, tier_label) for a Technician Totally Satisfied %."""
    if pct is None:
        return 0, 'N/A'
    for threshold, bonus in BONUS_TIERS:
        if pct >= threshold:
            return bonus, f'≥{threshold}%'
    return 0, '<92%'


def _totally_satisfied_pct(rows, field):
    """Calculate Totally Satisfied % from a list of survey rows."""
    total = sum(1 for r in rows if r.get(field))
    if total == 0:
        return None
    ts = sum(1 for r in rows if (r.get(field) or '').lower() == 'totally satisfied')
    return round(100 * ts / total)


def _load_ai_settings():
    try:
        with open(_SETTINGS_FILE) as f:
            settings = _json.load(f)
        cb = settings.get('chatbot', {})
        return cb.get('provider', ''), cb.get('api_key', ''), cb.get('primary_model', '')
    except Exception:
        return '', '', ''


def _call_openai(api_key, model, prompt):
    """Call OpenAI for executive summary."""
    try:
        resp = _requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": model or "gpt-4o-mini",
                "messages": [
                    {"role": "system", "content": "You are a fleet operations analyst for AAA roadside assistance. Write concise, actionable executive summaries. Use plain English. Be specific about which drivers and metrics. CRITICAL: Bonuses are per-driver based on individual TECHNICIAN satisfaction score (not overall). A driver with ≥92% tech score earns a bonus even if the garage average is below 92%. Never contradict the bonus data provided. Keep it under 200 words."},
                    {"role": "user", "content": prompt},
                ],
                "max_tokens": 512,
                "temperature": 0.3,
            },
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"]
    except Exception as e:
        log.warning('OpenAI call failed: %s', e)
        return None


@router.get("/api/garages/{territory_id}/performance-scorecard")
def api_garage_performance_scorecard(
    territory_id: str,
    start_date: str = Query(None, description="YYYY-MM-DD"),
    end_date: str = Query(None, description="YYYY-MM-DD"),
):
    """Garage performance scorecard: 4 satisfaction scores, primary/secondary split,
    driver breakdown with bonus calculation, AI executive summary."""

    territory_id = sanitize_soql(territory_id)

    # Default date range: current month
    today = _date.today()
    if not start_date:
        start_date = today.replace(day=1).isoformat()
    if not end_date:
        end_date = today.isoformat()

    cache_key = f'garage_perf_scorecard_{territory_id}_{start_date}_{end_date}'
    cached = cache.get(cache_key)
    if cached:
        return cached

    start_utc = f"{start_date}T00:00:00Z"
    end_utc = f"{end_date}T23:59:59Z"

    # ── Parallel queries: surveys + SAs + territory history ──
    def _get_surveys():
        return sf_query_all(f"""
            SELECT ERS_Work_Order__c,
                   ERS_Overall_Satisfaction__c,
                   ERS_Response_Time_Satisfaction__c,
                   ERS_Technician_Satisfaction__c,
                   ERSSatisfaction_With_Being_Kept_Informed__c,
                   ERS_Driver__r.Name,
                   Off_Platform_Driver__c,
                   Survey_Driver__c,
                   Customer_Comments__c,
                   ERS_Work_Order__r.WorkOrderNumber,
                   ERS_Work_Order__r.CreatedDate
            FROM Survey_Result__c
            WHERE ERS_Work_Order__r.ServiceTerritoryId = '{territory_id}'
              AND ERS_Work_Order__r.CreatedDate >= {start_utc}
              AND ERS_Work_Order__r.CreatedDate < {end_utc}
              AND ERS_Overall_Satisfaction__c != null
        """)

    def _get_sas():
        return sf_query_all(f"""
            SELECT Id, AppointmentNumber, ParentRecordId, CreatedDate, Status,
                   ActualStartTime, ERS_PTA__c,
                   WorkType.Name, ERS_Dispatch_Method__c,
                   ERS_Facility_Decline_Reason__c
            FROM ServiceAppointment
            WHERE ServiceTerritoryId = '{territory_id}'
              AND CreatedDate >= {start_utc}
              AND CreatedDate < {end_utc}
              AND RecordType.Name = 'ERS Service Appointment'
              AND WorkType.Name != 'Tow Drop-Off'
        """)

    def _get_territory_history():
        """First territory assignment for each SA — OldValue=null means original assignment.
        Field is 'ServiceTerritory', stores both name and ID in separate rows."""
        return sf_query_all(f"""
            SELECT ServiceAppointmentId, OldValue, NewValue, CreatedDate
            FROM ServiceAppointmentHistory
            WHERE Field = 'ServiceTerritory'
              AND ServiceAppointment.ServiceTerritoryId = '{territory_id}'
              AND ServiceAppointment.CreatedDate >= {start_utc}
              AND ServiceAppointment.CreatedDate < {end_utc}
              AND ServiceAppointment.RecordType.Name = 'ERS Service Appointment'
            ORDER BY ServiceAppointmentId, CreatedDate ASC
        """)

    def _get_towbook_on_location():
        """Real arrival for Towbook SAs — ActualStartTime is fake."""
        return sf_query_all(f"""
            SELECT ServiceAppointmentId, CreatedDate, NewValue
            FROM ServiceAppointmentHistory
            WHERE Field = 'Status'
              AND ServiceAppointment.ServiceTerritoryId = '{territory_id}'
              AND ServiceAppointment.CreatedDate >= {start_utc}
              AND ServiceAppointment.CreatedDate < {end_utc}
              AND ServiceAppointment.ERS_Dispatch_Method__c = 'Towbook'
              AND ServiceAppointment.Status = 'Completed'
              AND ServiceAppointment.RecordType.Name = 'ERS Service Appointment'
        """)

    data = sf_parallel(surveys=_get_surveys, sas=_get_sas,
                       territory_hist=_get_territory_history,
                       tb_on_loc=_get_towbook_on_location)
    surveys = data['surveys']
    sas = data['sas']
    territory_hist = data['territory_hist']

    # Build Towbook real arrival map: SA Id → On Location datetime
    towbook_arrival = {}
    for r in data['tb_on_loc']:
        if r.get('NewValue') != 'On Location':
            continue
        sa_id = r.get('ServiceAppointmentId')
        ts = _parse_dt(r.get('CreatedDate'))
        if ts and (sa_id not in towbook_arrival or ts < towbook_arrival[sa_id]):
            towbook_arrival[sa_id] = ts

    # Find SAs where this garage is NOT the original territory (secondary)
    # The first record per SA with OldValue=null is the original assignment.
    # If that NewValue != this territory_id, the SA cascaded here.
    original_territory = {}  # sa_id -> first territory ID
    for row in territory_hist:
        sa_id = row.get('ServiceAppointmentId')
        if sa_id in original_territory:
            continue  # already have the first record
        if row.get('OldValue') is None:
            nv = row.get('NewValue') or ''
            # ID rows are 15 or 18 chars starting with 0H
            if len(nv) >= 15 and nv.startswith('0H'):
                original_territory[sa_id] = nv

    reassigned_sa_ids = {sa_id for sa_id, orig_tid in original_territory.items()
                         if orig_tid != territory_id}

    total_completed = sum(1 for sa in sas if sa.get('Status') == 'Completed')

    # ── Classify SAs as primary vs secondary ──
    # Build SA lookup by ID (for territory classification)
    sa_by_id = {sa['Id']: sa for sa in sas}

    primary_sas = [sa for sa in sas if sa['Id'] not in reassigned_sa_ids]
    secondary_sas = [sa for sa in sas if sa['Id'] in reassigned_sa_ids]

    def _sa_stats(sa_list):
        total = len(sa_list)
        completed = sum(1 for sa in sa_list if sa.get('Status') == 'Completed')
        cancelled = sum(1 for sa in sa_list if 'cancel' in (sa.get('Status') or '').lower())
        declined = sum(1 for sa in sa_list if sa.get('ERS_Facility_Decline_Reason__c'))
        # ATA: Fleet = ActualStartTime, Towbook = SAHistory "On Location"
        ata_vals = []
        pta_vals = []
        for sa in sa_list:
            if sa.get('Status') != 'Completed':
                continue
            created = _parse_dt(sa.get('CreatedDate'))
            dm = sa.get('ERS_Dispatch_Method__c') or ''
            if dm == 'Towbook':
                actual = towbook_arrival.get(sa['Id'])
            else:
                actual = _parse_dt(sa.get('ActualStartTime'))
            if created and actual:
                ata = (actual - created).total_seconds() / 60
                if 0 < ata < 480:
                    ata_vals.append(ata)
            pta = sa.get('ERS_PTA__c')
            if pta and 0 < float(pta) < 999:
                pta_vals.append(float(pta))
        avg_ata = round(sum(ata_vals) / len(ata_vals)) if ata_vals else None
        avg_pta = round(sum(pta_vals) / len(pta_vals)) if pta_vals else None
        pta_hit = sum(1 for a, p in zip(ata_vals, pta_vals) if a <= p) if ata_vals and pta_vals else None
        pta_hit_pct = round(100 * pta_hit / len(ata_vals)) if pta_hit is not None and ata_vals else None
        return {
            'total_sas': total,
            'completed': completed,
            'cancelled': cancelled,
            'declined': declined,
            'avg_ata': avg_ata,
            'avg_pta': avg_pta,
            'pta_hit_pct': pta_hit_pct,
        }

    # ── Map WorkOrder → SA to classify surveys as primary/secondary ──
    # SA.ParentRecordId = WOLI.Id. Query WOLI to get WorkOrderId.
    woli_ids = [sa.get('ParentRecordId') for sa in sas if sa.get('ParentRecordId')]
    wo_to_sa = {}  # WO Id → SA Id
    wo_to_sa_number = {}  # WO Id → SA AppointmentNumber
    if woli_ids:
        # Batch query WOLI in chunks of 200
        for i in range(0, len(woli_ids), 200):
            chunk = woli_ids[i:i+200]
            id_list = "','".join(chunk)
            woli_rows = sf_query_all(f"SELECT Id, WorkOrderId FROM WorkOrderLineItem WHERE Id IN ('{id_list}')")
            woli_to_wo = {r['Id']: r.get('WorkOrderId') for r in woli_rows}
            for sa in sas:
                woli_id = sa.get('ParentRecordId')
                wo_id = woli_to_wo.get(woli_id)
                if wo_id and wo_id not in wo_to_sa:
                    wo_to_sa[wo_id] = sa['Id']
                    wo_to_sa_number[wo_id] = sa.get('AppointmentNumber', '')

    primary_surveys = []
    secondary_surveys = []
    for sv in surveys:
        wo_id = sv.get('ERS_Work_Order__c') or ''
        sa_id = wo_to_sa.get(wo_id)
        if sa_id and sa_id in reassigned_sa_ids:
            secondary_surveys.append(sv)
        else:
            primary_surveys.append(sv)

    # ── Garage-level scores ──
    overall_pct = _totally_satisfied_pct(surveys, 'ERS_Overall_Satisfaction__c')
    rt_pct = _totally_satisfied_pct(surveys, 'ERS_Response_Time_Satisfaction__c')
    tech_pct = _totally_satisfied_pct(surveys, 'ERS_Technician_Satisfaction__c')
    informed_pct = _totally_satisfied_pct(surveys, 'ERSSatisfaction_With_Being_Kept_Informed__c')

    bonus_per_sa, bonus_tier = _bonus_for_pct(tech_pct)
    total_bonus = bonus_per_sa * total_completed

    # ── Primary vs Secondary scores + SA stats ──
    primary_sa_stats = _sa_stats(primary_sas)
    secondary_sa_stats = _sa_stats(secondary_sas)

    def _group_scores(survey_group, sa_stats):
        return {
            'overall_pct': _totally_satisfied_pct(survey_group, 'ERS_Overall_Satisfaction__c'),
            'response_time_pct': _totally_satisfied_pct(survey_group, 'ERS_Response_Time_Satisfaction__c'),
            'technician_pct': _totally_satisfied_pct(survey_group, 'ERS_Technician_Satisfaction__c'),
            'kept_informed_pct': _totally_satisfied_pct(survey_group, 'ERSSatisfaction_With_Being_Kept_Informed__c'),
            'survey_count': len(survey_group),
            **sa_stats,
        }

    # ── Driver breakdown ──
    driver_map = defaultdict(list)  # driver_name -> [survey rows]
    for sv in surveys:
        # Driver name: Fleet = ERS_Driver__r.Name, Towbook = Survey_Driver__c (formula)
        driver = (sv.get('ERS_Driver__r') or {}).get('Name') or sv.get('Survey_Driver__c') or 'Unknown'
        driver_map[driver].append(sv)

    drivers = []
    for name, svs in sorted(driver_map.items(), key=lambda x: len(x[1]), reverse=True):
        d_tech_pct = _totally_satisfied_pct(svs, 'ERS_Technician_Satisfaction__c')
        d_bonus_per_sa, d_bonus_tier = _bonus_for_pct(d_tech_pct)
        d_sa_count = len(svs)  # approximate: survey count ≈ SA count for this driver
        d_total_bonus = d_bonus_per_sa * d_sa_count

        # Individual survey details for drill-down
        survey_details = []
        for sv in svs:
            wo = sv.get('ERS_Work_Order__r') or {}
            created = _parse_dt(wo.get('CreatedDate'))
            wo_id = sv.get('ERS_Work_Order__c') or ''
            sa_num = wo_to_sa_number.get(wo_id, '')
            survey_details.append({
                'wo_number': wo.get('WorkOrderNumber', ''),
                'sa_number': sa_num,
                'call_date': created.strftime('%Y-%m-%d') if created else '',
                'overall': sv.get('ERS_Overall_Satisfaction__c'),
                'response_time': sv.get('ERS_Response_Time_Satisfaction__c'),
                'technician': sv.get('ERS_Technician_Satisfaction__c'),
                'kept_informed': sv.get('ERSSatisfaction_With_Being_Kept_Informed__c'),
                'comment': (sv.get('Customer_Comments__c') or '').strip() or None,
            })

        drivers.append({
            'name': name,
            'survey_count': len(svs),
            'overall_pct': _totally_satisfied_pct(svs, 'ERS_Overall_Satisfaction__c'),
            'response_time_pct': _totally_satisfied_pct(svs, 'ERS_Response_Time_Satisfaction__c'),
            'technician_pct': d_tech_pct,
            'kept_informed_pct': _totally_satisfied_pct(svs, 'ERSSatisfaction_With_Being_Kept_Informed__c'),
            'bonus_per_sa': d_bonus_per_sa,
            'bonus_tier': d_bonus_tier,
            'total_bonus': d_total_bonus,
            'surveys': survey_details,
        })

    result = {
        'territory_id': territory_id,
        'start_date': start_date,
        'end_date': end_date,
        'garage_summary': {
            'overall_pct': overall_pct,
            'response_time_pct': rt_pct,
            'technician_pct': tech_pct,
            'kept_informed_pct': informed_pct,
            'total_surveys': len(surveys),
            'total_sas': len(sas),
            'total_completed': total_completed,
            'bonus_tier': bonus_tier,
            'bonus_per_sa': bonus_per_sa,
            'total_bonus': total_bonus,
        },
        'primary_vs_secondary': {
            'primary': _group_scores(primary_surveys, primary_sa_stats),
            'secondary': _group_scores(secondary_surveys, secondary_sa_stats),
        },
        'drivers': drivers,
        'ai_summary': None,  # loaded async via separate endpoint
    }

    cache.put(cache_key, result, ttl=3600)
    return result


@router.get("/api/garages/{territory_id}/performance-scorecard/ai-summary")
def api_garage_ai_summary(
    territory_id: str,
    start_date: str = Query(None),
    end_date: str = Query(None),
):
    """Generate AI executive summary for garage performance."""
    territory_id = sanitize_soql(territory_id)
    today = _date.today()
    if not start_date:
        start_date = today.replace(day=1).isoformat()
    if not end_date:
        end_date = today.isoformat()

    cache_key = f'garage_ai_summary_v2_{territory_id}_{start_date}_{end_date}'
    cached = cache.get(cache_key)
    if cached:
        return cached

    # Get the scorecard data first
    scorecard = api_garage_performance_scorecard(territory_id, start_date, end_date)
    gs = scorecard['garage_summary']
    drivers = scorecard['drivers']
    ps_p = scorecard['primary_vs_secondary']['primary']
    ps_s = scorecard['primary_vs_secondary']['secondary']

    if not drivers:
        return {'summary': 'No survey data available for this period.'}

    # Build prompt
    driver_lines = []
    for d in drivers[:15]:  # top 15 drivers
        driver_lines.append(
            f"  {d['name']}: {d['survey_count']} surveys, "
            f"Overall={d['overall_pct']}%, Tech={d['technician_pct']}%, "
            f"ResponseTime={d['response_time_pct']}%, Informed={d['kept_informed_pct']}%, "
            f"Bonus=${d['total_bonus']}"
        )

    # Compute total driver bonuses for accuracy
    total_driver_bonuses = sum(d.get('total_bonus', 0) for d in drivers)
    drivers_with_bonus = sum(1 for d in drivers if d.get('total_bonus', 0) > 0)
    drivers_without_bonus = sum(1 for d in drivers if d.get('total_bonus', 0) == 0 and d.get('technician_pct') is not None)

    prompt = f"""Analyze this garage performance scorecard for {start_date} to {end_date}:

GARAGE SCORES (Totally Satisfied %):
- Overall: {gs['overall_pct']}% ({gs['total_surveys']} surveys, {gs['total_completed']} completed SAs)
- Response Time: {gs['response_time_pct']}%
- Technician: {gs['technician_pct']}%
- Kept Informed: {gs['kept_informed_pct']}%

BONUS RULES (IMPORTANT — read carefully):
- Bonuses are based ONLY on TECHNICIAN satisfaction score, NOT overall score.
- Bonuses are calculated PER DRIVER based on each driver's INDIVIDUAL technician score.
- Tiers: ≥98% = $4/SA, ≥96% = $3/SA, ≥94% = $2/SA, ≥92% = $1/SA, <92% = $0.
- Garage-wide bonus: {gs['bonus_tier']} tier → ${gs['bonus_per_sa']}/SA × {gs['total_completed']} SAs = ${gs['total_bonus']} total
- Individual driver totals: {drivers_with_bonus} drivers earned bonuses totaling ${total_driver_bonuses:,.0f}. {drivers_without_bonus} drivers earned $0.
- A driver can earn a bonus even if the garage average is below 92%, as long as their individual tech score is ≥92%.

PRIMARY vs SECONDARY:
- Primary (first assigned): {ps_p.get('total_sas',0)} SAs, {ps_p.get('completed',0)} completed, {ps_p.get('declined',0)} declined, ATA={ps_p.get('avg_ata','N/A')}m, PTA hit={ps_p.get('pta_hit_pct','N/A')}%, {ps_p.get('survey_count',0)} surveys, Tech={ps_p.get('technician_pct')}%
- Secondary (reassigned here): {ps_s.get('total_sas',0)} SAs, {ps_s.get('completed',0)} completed, {ps_s.get('declined',0)} declined, ATA={ps_s.get('avg_ata','N/A')}m, PTA hit={ps_s.get('pta_hit_pct','N/A')}%, {ps_s.get('survey_count',0)} surveys, Tech={ps_s.get('technician_pct')}%

DRIVER BREAKDOWN:
{chr(10).join(driver_lines)}

Write a 3-4 paragraph executive summary:
1. Overall performance assessment — cite the 4 satisfaction scores. Note which are strong and which need improvement.
2. Bonus analysis — state exactly how many drivers earned bonuses and the total dollar amount. Name the top earners with their tech scores and bonus amounts. Do NOT say "no bonuses" if individual drivers earned bonuses.
3. Which drivers are underperforming (tech score <92%) and dragging the garage average down? Name them with scores.
4. Specific action items to improve — focus on the weakest satisfaction categories (response time, kept informed, etc.)."""

    provider, api_key, model = _load_ai_settings()
    if not api_key:
        return {'summary': 'AI not configured. Go to Admin → AI Assistant to set up.'}

    summary = _call_openai(api_key, model, prompt)
    result = {'summary': summary or 'Failed to generate summary.'}
    cache.put(cache_key, result, ttl=7200)
    return result


@router.get("/api/garages/{territory_id}/performance-scorecard/export")
def api_garage_export(
    territory_id: str,
    start_date: str = Query(None),
    end_date: str = Query(None),
):
    """Export garage performance scorecard to Excel."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse

    territory_id = sanitize_soql(territory_id)
    today = _date.today()
    if not start_date:
        start_date = today.replace(day=1).isoformat()
    if not end_date:
        end_date = today.isoformat()

    # Get scorecard data (will use cache if available)
    scorecard = api_garage_performance_scorecard(territory_id, start_date, end_date)
    gs = scorecard['garage_summary']
    ps = scorecard['primary_vs_secondary']
    drivers = scorecard['drivers']

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    amber_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    def _score_fill(pct):
        if pct is None:
            return None
        if pct >= 92:
            return green_fill
        if pct >= 82:
            return None
        if pct >= 70:
            return amber_fill
        return red_fill

    def _style_header(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = 'Garage Summary'
    ws.append(['Garage Performance Scorecard'])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Period: {start_date} to {end_date}'])
    ws.append([])
    ws.append(['Metric', 'Value'])
    _style_header(ws, 4, 2)
    ws.append(['Total Surveys', gs.get('total_surveys', 0)])
    ws.append(['Total Completed SAs', gs.get('total_completed', 0)])
    ws.append(['Overall Satisfaction %', gs.get('overall_pct')])
    ws.append(['Response Time Satisfaction %', gs.get('response_time_pct')])
    ws.append(['Technician Satisfaction %', gs.get('technician_pct')])
    ws.append(['Kept Informed Satisfaction %', gs.get('kept_informed_pct')])
    ws.append(['Bonus Tier', gs.get('bonus_tier', 'N/A')])
    ws.append(['Bonus per SA', f"${gs.get('bonus_per_sa', 0)}"])
    ws.append(['Total Garage Bonus', f"${gs.get('total_bonus', 0):,}"])
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18

    # Color code score cells
    for row_num in range(7, 11):
        cell = ws.cell(row=row_num, column=2)
        fill = _score_fill(cell.value)
        if fill:
            cell.fill = fill

    # ── Sheet 2: Driver Breakdown ──
    ws2 = wb.create_sheet('Drivers')
    headers = ['Driver', 'Surveys', 'Overall %', 'Response Time %', 'Technician %',
               'Kept Informed %', 'Bonus Tier', '$/SA', 'Total Bonus']
    ws2.append(headers)
    _style_header(ws2, 1, len(headers))

    for d in drivers:
        row = [
            d['name'], d['survey_count'],
            d.get('overall_pct'), d.get('response_time_pct'),
            d.get('technician_pct'), d.get('kept_informed_pct'),
            d.get('bonus_tier', 'N/A'), d.get('bonus_per_sa', 0),
            d.get('total_bonus', 0),
        ]
        ws2.append(row)
        row_num = ws2.max_row
        # Color code scores
        for col_idx, val in [(3, d.get('overall_pct')), (4, d.get('response_time_pct')),
                              (5, d.get('technician_pct')), (6, d.get('kept_informed_pct'))]:
            fill = _score_fill(val)
            if fill:
                ws2.cell(row=row_num, column=col_idx).fill = fill
        # Border all cells
        for col in range(1, len(headers) + 1):
            ws2.cell(row=row_num, column=col).border = thin_border

    # Auto-width
    for col_idx, _ in enumerate(headers, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 16
    ws2.column_dimensions['A'].width = 28

    # ── Sheet 3: Primary vs Secondary ──
    ws3 = wb.create_sheet('Primary vs Secondary')
    ps_headers = ['Metric', 'Primary', 'Secondary']
    ws3.append(ps_headers)
    _style_header(ws3, 1, 3)
    p = ps.get('primary', {})
    s = ps.get('secondary', {})
    for label, key in [('Total SAs', 'total_sas'), ('Completed', 'completed'),
                        ('Declined', 'declined'), ('Avg ATA (min)', 'avg_ata'),
                        ('PTA Hit %', 'pta_hit_pct'), ('Surveys', 'survey_count'),
                        ('Overall %', 'overall_pct'), ('Response Time %', 'response_time_pct'),
                        ('Technician %', 'technician_pct'), ('Kept Informed %', 'kept_informed_pct')]:
        ws3.append([label, p.get(key), s.get(key)])
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 14

    # ── Sheet 4: All Survey Details ──
    ws4 = wb.create_sheet('Survey Details')
    sv_headers = ['Driver', 'WO Number', 'Call Date', 'Overall', 'Response Time',
                  'Technician', 'Kept Informed', 'Comment']
    ws4.append(sv_headers)
    _style_header(ws4, 1, len(sv_headers))

    for d in drivers:
        for sv in d.get('surveys', []):
            ws4.append([
                d['name'], sv.get('wo_number', ''), sv.get('call_date', ''),
                sv.get('overall', ''), sv.get('response_time', ''),
                sv.get('technician', ''), sv.get('kept_informed', ''),
                sv.get('comment', ''),
            ])

    ws4.column_dimensions['A'].width = 28
    ws4.column_dimensions['B'].width = 14
    ws4.column_dimensions['C'].width = 12
    for c in ['D', 'E', 'F', 'G']:
        ws4.column_dimensions[c].width = 18
    ws4.column_dimensions['H'].width = 50

    # ── Sheet 5: Flat SA Data (analyst verification) ──
    # Query all SAs with full details + territory history for proof
    start_utc = f"{start_date}T00:00:00Z"
    end_utc = f"{end_date}T23:59:59Z"
    territory_id = sanitize_soql(territory_id)

    sa_detail = sf_query_all(f"""
        SELECT Id, AppointmentNumber, CreatedDate, Status,
               ActualStartTime, ERS_PTA__c, WorkType.Name,
               ERS_Dispatch_Method__c, ERS_Facility_Decline_Reason__c,
               ServiceTerritory.Name, ParentRecordId
        FROM ServiceAppointment
        WHERE ServiceTerritoryId = '{territory_id}'
          AND CreatedDate >= {start_utc}
          AND CreatedDate < {end_utc}
          AND RecordType.Name = 'ERS Service Appointment'
          AND WorkType.Name != 'Tow Drop-Off'
    """)

    # Get territory history for primary/secondary proof
    territory_hist = sf_query_all(f"""
        SELECT ServiceAppointmentId, OldValue, NewValue, CreatedDate
        FROM ServiceAppointmentHistory
        WHERE Field = 'ServiceTerritory'
          AND ServiceAppointment.ServiceTerritoryId = '{territory_id}'
          AND ServiceAppointment.CreatedDate >= {start_utc}
          AND ServiceAppointment.CreatedDate < {end_utc}
          AND ServiceAppointment.RecordType.Name = 'ERS Service Appointment'
        ORDER BY ServiceAppointmentId, CreatedDate ASC
    """)

    # Build original territory map
    original_map = {}  # sa_id -> (original_tid, original_name)
    for row in territory_hist:
        sa_id = row['ServiceAppointmentId']
        if sa_id in original_map:
            continue
        if row.get('OldValue') is None:
            nv = row.get('NewValue') or ''
            if len(nv) >= 15 and nv.startswith('0H'):
                original_map[sa_id] = nv
            elif nv and not nv.startswith('0H'):
                # This is the name row — store as name
                if sa_id not in original_map:
                    original_map[sa_id] = nv  # will be overwritten by ID if ID comes next

    # Build name lookup from territory history
    original_names = {}
    for row in territory_hist:
        sa_id = row['ServiceAppointmentId']
        if sa_id in original_names:
            continue
        if row.get('OldValue') is None:
            nv = row.get('NewValue') or ''
            if nv and not nv.startswith('0H'):
                original_names[sa_id] = nv

    # Get assigned driver for each SA
    sa_ids_str = "','".join(sa['Id'] for sa in sa_detail[:500])
    assigned_drivers = {}
    if sa_ids_str:
        for chunk_start in range(0, len(sa_detail), 200):
            chunk = sa_detail[chunk_start:chunk_start+200]
            chunk_ids = "','".join(sa['Id'] for sa in chunk)
            ar_rows = sf_query_all(f"""
                SELECT ServiceAppointmentId, ServiceResource.Name
                FROM AssignedResource
                WHERE ServiceAppointmentId IN ('{chunk_ids}')
            """)
            for ar in ar_rows:
                assigned_drivers[ar['ServiceAppointmentId']] = (ar.get('ServiceResource') or {}).get('Name', '')

    # Build survey lookup by WO Id for joining
    survey_by_wo = {}
    woli_ids = [sa.get('ParentRecordId') for sa in sa_detail if sa.get('ParentRecordId')]
    wo_to_sa_export = {}
    if woli_ids:
        for i in range(0, len(woli_ids), 200):
            chunk = woli_ids[i:i+200]
            id_list = "','".join(chunk)
            woli_rows = sf_query_all(f"SELECT Id, WorkOrderId FROM WorkOrderLineItem WHERE Id IN ('{id_list}')")
            for r in woli_rows:
                for sa in sa_detail:
                    if sa.get('ParentRecordId') == r['Id']:
                        wo_to_sa_export[r['WorkOrderId']] = sa['Id']
                        break
    sa_to_wo = {v: k for k, v in wo_to_sa_export.items()}

    # Get all surveys for this garage/period
    all_surveys = sf_query_all(f"""
        SELECT ERS_Work_Order__c, ERS_Overall_Satisfaction__c,
               ERS_Response_Time_Satisfaction__c, ERS_Technician_Satisfaction__c,
               ERSSatisfaction_With_Being_Kept_Informed__c,
               Customer_Comments__c, ERS_NPS__c, ERS_Renew__c
        FROM Survey_Result__c
        WHERE ERS_Work_Order__r.ServiceTerritoryId = '{territory_id}'
          AND ERS_Work_Order__r.CreatedDate >= {start_utc}
          AND ERS_Work_Order__r.CreatedDate < {end_utc}
          AND ERS_Overall_Satisfaction__c != null
    """)
    # Map WO Id -> survey
    for sv in all_surveys:
        wo_id = sv.get('ERS_Work_Order__c')
        if wo_id:
            sa_id = wo_to_sa_export.get(wo_id)
            if sa_id:
                survey_by_wo[sa_id] = sv

    ws5 = wb.create_sheet('SA Flat Data')
    sa_headers = [
        'SA Number', 'SA Id', 'Created Date', 'Status', 'Work Type',
        'Dispatch Method', 'Current Territory', 'Original Territory',
        'Classification', 'Assigned Driver', 'Decline Reason',
        'ATA (min)', 'PTA (min)', 'PTA Hit',
        'Survey Overall', 'Survey Response Time', 'Survey Technician',
        'Survey Kept Informed', 'NPS', 'Renewal', 'Customer Comment',
    ]
    ws5.append(sa_headers)
    _style_header(ws5, 1, len(sa_headers))

    for sa in sa_detail:
        sa_id = sa['Id']
        created = _parse_dt(sa.get('CreatedDate'))
        actual = _parse_dt(sa.get('ActualStartTime'))
        ata = None
        if created and actual:
            ata_val = (actual - created).total_seconds() / 60
            if 0 < ata_val < 480:
                ata = round(ata_val)
        pta = sa.get('ERS_PTA__c')
        pta_val = float(pta) if pta and 0 < float(pta) < 999 else None
        pta_hit = 'Yes' if ata and pta_val and ata <= pta_val else ('No' if ata and pta_val else '')

        orig_name = original_names.get(sa_id, '')
        orig_id = original_map.get(sa_id, '')
        # Classification
        if orig_id and len(orig_id) >= 15 and orig_id.startswith('0H'):
            classification = 'Primary' if orig_id == territory_id else 'Secondary'
        else:
            classification = 'Primary'  # no history = was always here

        current_territory = (sa.get('ServiceTerritory') or {}).get('Name', '')
        driver = assigned_drivers.get(sa_id, '')

        # Join survey data
        sv = survey_by_wo.get(sa_id, {})

        ws5.append([
            sa.get('AppointmentNumber', ''),
            sa_id,
            created.strftime('%Y-%m-%d %H:%M') if created else '',
            sa.get('Status', ''),
            (sa.get('WorkType') or {}).get('Name', ''),
            sa.get('ERS_Dispatch_Method__c', ''),
            current_territory,
            orig_name or '',
            classification,
            driver,
            sa.get('ERS_Facility_Decline_Reason__c', '') or '',
            ata,
            round(pta_val) if pta_val else '',
            pta_hit,
            sv.get('ERS_Overall_Satisfaction__c', ''),
            sv.get('ERS_Response_Time_Satisfaction__c', ''),
            sv.get('ERS_Technician_Satisfaction__c', ''),
            sv.get('ERSSatisfaction_With_Being_Kept_Informed__c', ''),
            sv.get('ERS_NPS__c', ''),
            sv.get('ERS_Renew__c', ''),
            (sv.get('Customer_Comments__c') or '').strip() if sv else '',
        ])
        row_num = ws5.max_row
        # Color classification
        cls_cell = ws5.cell(row=row_num, column=9)
        if classification == 'Secondary':
            cls_cell.fill = amber_fill
        # Border
        for col in range(1, len(sa_headers) + 1):
            ws5.cell(row=row_num, column=col).border = thin_border

    # Auto-width
    widths = [14, 20, 18, 12, 16, 16, 30, 30, 12, 24, 30, 10, 10, 8,
              20, 20, 20, 20, 6, 8, 50]
    for i, w in enumerate(widths, 1):
        ws5.column_dimensions[ws5.cell(row=1, column=i).column_letter].width = w

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"garage_scorecard_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
